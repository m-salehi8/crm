import os
import django

# =========================
# تنظیمات Django
# =========================
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "backend.settings")
django.setup()

# =========================
# ایمپورت‌ها
# =========================
from openpyxl import load_workbook
from core.models import User
from hr.models import DeductionType, DeductionWork

# =========================
# مسیر فایل اکسل
# =========================
file_path = "fn/fn.xlsx"

# =========================
# باز کردن فایل اکسل
# =========================
wb = load_workbook(file_path)
ws = wb.active

# =========================
# دریافت DeductionType ها (یک بار)
# =========================
DEDUCTION_TYPES = {
    "tak": DeductionType.objects.get(key="tak"),
    "vam": DeductionType.objects.get(key="vam"),
    "san": DeductionType.objects.get(key="san"),
    "food": DeductionType.objects.get(key="food"),
    "digipay": DeductionType.objects.get(key="digipay"),
    "mashhad": DeductionType.objects.get(key="mashhad"),
}

# =========================
# رد کردن هدر
# =========================
rows = ws.iter_rows(values_only=True)
next(rows)

# =========================
# پردازش داده‌ها
# =========================
for row in rows:
    personnel_code = row[0]
    year = row[3]
    month = row[4]

    user = User.objects.filter(personnel_code=personnel_code).first()
    if not user:
        print(f"❌ User not found: {personnel_code}")
        continue

    deductions_data = {
        "tak": row[5],
        "vam": row[6],
        # در صورت نیاز ستون‌های بیشتر اضافه کن
    }

    for key, value in deductions_data.items():
        if not value:
            continue

        deduction_type = DEDUCTION_TYPES[key]

        deduction, created = DeductionWork.objects.get_or_create(
            user=user,
            year=year,
            month=month,
            type=deduction_type,
        )

        deduction.value = value
        deduction.save()

        status = "CREATED" if created else "UPDATED"
        print(f"✅ {status} | {user.id} | {key} | {value}")

print("🎉 Done")
