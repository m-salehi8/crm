import jdatetime
from rest_framework.permissions import IsAuthenticated
from weasyprint import HTML
from core.models import User
from core.permissions import *
from rest_framework import status
from django.db.models import Q, Count, Sum, Case, When, BooleanField
from django.http import HttpResponse
from core.serializers import SerUserList
from rest_framework.response import Response
from django.template.loader import render_to_string
from .models import InvoiceCover, InvoiceCategory, Invoice
from .serializers import SerInvoiceCover, SerInvoiceCategory, SerInvoice, SerInvoiceCoverTask
from rest_framework.generics import GenericAPIView, ListAPIView, get_object_or_404, DestroyAPIView, RetrieveAPIView
from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class InvoiceCoverList(GenericAPIView):
    permission_classes = [IsInvoiceCollector]

    def get(self, request):
        if request.user.groups.filter(name__in=['invoice-confirm1', 'invoice-confirm3', 'invoice-accept', 'invoice-deposit']).exists():  # کارشناس مالی، معاون برنامه‌ریزی، مدیرکل مالی، کارشناس پرداخت
            covers = InvoiceCover.objects.all()
        elif request.user.is_head_of_unit:
            covers = InvoiceCover.objects.filter(user__post__unit=request.user.post.unit)  # مدیر واحد
        else:
            covers = InvoiceCover.objects.filter(user=request.user)
        user = int(request.GET.get('user', 0))
        if user:
            covers = covers.filter(user_id=user)
        _type = request.GET.get('type', 'همه')
        if _type != 'همه':
            covers = covers.filter(type=_type)
        locked = request.GET.get('locked', 'همه')
        if locked != 'همه':
            covers = covers.filter(locked=locked == 'بله')
        confirm1 = request.GET.get('confirm1', 'همه')
        if confirm1 != 'همه':
            covers = covers.filter(confirm1=True if confirm1 == 'بله' else False if confirm1 == 'خیر' else None)
        confirm2 = request.GET.get('confirm2', 'همه')
        if confirm2 != 'همه':
            covers = covers.filter(confirm2=True if confirm2 == 'بله' else False if confirm2 == 'خیر' else None)
        confirm3 = request.GET.get('confirm3', 'همه')
        if confirm3 != 'همه':
            covers = covers.filter(confirm3=True if confirm3 == 'بله' else False if confirm3 == 'خیر' else None)
        accepted = request.GET.get('accepted', 'همه')
        if accepted != 'همه':
            covers = covers.filter(accepted=True if accepted == 'بله' else False if accepted == 'خیر' else None)
        deposit = request.GET.get('deposit', 'همه')
        if deposit != 'همه':
            covers = covers.filter(deposit_time__isnull=deposit == 'خیر')
        search = request.GET.get('search', '')
        if search:
            covers = covers.filter(Q(no__contains=search) | Q(invoices__description__contains=search)).distinct()
        if request.GET.get('need_action', 'همه') == 'منتظر اقدام':
            _ = covers.filter(Q(user=request.user) & (Q(confirm1=False) | Q(confirm2=False) | Q(confirm3=False) | Q(accepted=False)))
            if request.user.groups.filter(name='invoice-confirm1').exists():
                _ = _ | covers.filter(locked=True, confirm1=None)
            if request.user.is_head_of_unit:
                _ = _ | covers.filter(unit=request.user.post.unit, confirm1=True, confirm2=None)
            if request.user.groups.filter(name='invoice-confirm3').exists():
                _ = _ | covers.filter(confirm2=True, confirm3=None)
            if request.user.groups.filter(name='invoice-accept').exists():
                _ = _ | covers.filter(confirm3=True, accepted=None)
            if request.user.groups.filter(name='invoice-deposit').exists():
                _ = _ | covers.filter(type='مستقیم', accepted=True, deposit_time=None)
            covers = _.distinct()
        size = request.user.profile.page_size
        page = min(int(request.GET.get('page', 1)), covers.count() // size + 1)
        data = {
            'count': covers.count(),
            'sum': Invoice.objects.filter(cover__in=covers).aggregate(val=Sum('price'))['val'] or 0,
            'page': page,
            'can_edit':  request.user.groups.filter(name__in=['invoice-confirm1']).exists(),
            'list': SerInvoiceCover(covers[size*(page-1):size*page], many=True, context={'user': request.user}).data
        }
        return Response(data=data)


class InvoiceCategoryList(ListAPIView):
    queryset = InvoiceCategory.objects.filter(is_available=True)
    serializer_class = SerInvoiceCategory


class InvoiceCoverDetail(ListAPIView):
    permission_classes = [IsInvoiceCollector]
    serializer_class = SerInvoice

    def get_queryset(self):
        return Invoice.objects.filter(cover_id=self.kwargs['pk'])



class GetInvoiceCoverPdf(GenericAPIView):
    def get(self, request, pk):
        cover = get_object_or_404(InvoiceCover, pk=pk)
        user = request.user
        if cover.user == user or (user.is_head_of_unit and user.post.unit in [cover.unit, cover.unit.parent]) or request.user.groups.filter(name__in=['invoice-registrar', 'invoice-confirm1', 'invoice-confirm3', 'invoice-accept', 'invoice-deposit']).exists():
            html_string = render_to_string('invoice_list.html', {'cover': cover})
            response = HttpResponse(HTML(string=html_string).write_pdf(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="cover{cover.id}.pdf"'
            return response
        return HttpResponse(status=403)



class GetInvoiceCoverExcel(GenericAPIView):
    def get(self, request, pk):
        cover = get_object_or_404(InvoiceCover, pk=pk)
        user = request.user
        if cover.user == user or (user.is_head_of_unit and user.post.unit in [cover.unit, cover.unit.parent]) or request.user.groups.filter(name__in=['invoice-registrar', 'invoice-confirm1', 'invoice-confirm3', 'invoice-accept', 'invoice-deposit']).exists():
            # Create a workbook and add a worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = f"روکش فاکتور {cover.id}"

            # Define styles
            header_font = Font(name='Tahoma', size=12, bold=True)
            header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            regular_font = Font(name='Tahoma', size=10)
            regular_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Add main title
            ws.merge_cells('A1:K2')
            title_cell = ws['A1']
            cover_type = 'تنخواه' if cover.type == 'عادی' else 'سند مستقیم'
            status_text = ''
            if cover.locked:
                status = 'در حال بررسی'
                if cover.accepted:
                    status = 'تأیید شده'
                elif cover.accepted is False:
                    status = 'عدم تأیید'
                status_text = f'، {status}'
            else:
                status_text = '، درحال ویرایش'

            title_cell.value = f'روکش {cover_type} شماره {cover.no} ({status_text})'
            title_cell.font = Font(name='Tahoma', size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 30

            # Add summary information
            ws.merge_cells('A3:B4')
            summary_cell = ws['A3']
            total_sum = cover.invoices.aggregate(val=Sum('price'))['val'] or 0
            summary_cell.value = f'تعداد سند: {cover.invoice_count} ردیف\nجمع مبلغ: {total_sum:,} ریال'
            summary_cell.font = regular_font
            summary_cell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells('C3:D4')
            unit_cell = ws['C3']
            unit_cell.value = f'واحد: {cover.unit.title}\nتاریخ: از {cover.begin_date} تا {cover.end_date}'
            unit_cell.font = regular_font
            unit_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Add header row for invoice details
            headers = ['ردیف', 'شماره', 'مرکز هزینه', 'تاریخ', 'ردیف معین', 'مبلغ', 'ارزش افزوده', 'مبلغ کل', 'فاکتور', 'طرف مقابل', 'شرح']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Add invoice data
            row_num = 6
            for invoice in cover.invoices.all():
                has_paper = '✔' if invoice.has_paper else ''
                # Calculate gross price and value added price without formatting
                gross_price = (round(invoice.price / 1.1) if invoice.value_added else invoice.price)
                value_added_price = (round(invoice.price / 11) if invoice.value_added else 0)

                row = [
                    row_num - 5,  # ردیف
                    invoice.no or '',  # شماره
                    invoice.unit.title if invoice.unit else 'مرکز هزینه مشترک',  # مرکز هزینه
                    str(invoice.date) if invoice.date else '',  # تاریخ
                    invoice.category.title if invoice.category else '',  # ردیف معین
                    gross_price,  # مبلغ
                    value_added_price,  # ارزش افزوده
                    invoice.price,  # مبلغ کل
                    has_paper,  # فاکتور
                    invoice.issuer or '',  # طرف مقابل
                    invoice.description or '',  # شرح
                ]

                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.font = regular_font
                    cell.alignment = regular_alignment
                    cell.border = border
                row_num += 1

            # Add totals row
            ws.merge_cells(f'A{row_num}:E{row_num}')
            total_label_cell = ws[f'A{row_num}']
            total_label_cell.value = 'جمع'
            total_label_cell.font = header_font
            total_label_cell.alignment = header_alignment
            total_label_cell.border = border

            # Calculate raw numeric totals for the Excel sheet
            price_sum = sum((round(inv.price / 1.1) if inv.value_added else inv.price) for inv in cover.invoices.all())
            value_added_sum = sum((round(inv.price / 11) if inv.value_added else 0) for inv in cover.invoices.all())
            total_price_sum = sum(inv.price for inv in cover.invoices.all())

            ws[f'F{row_num}'].value = price_sum
            ws[f'F{row_num}'].font = header_font
            ws[f'F{row_num}'].alignment = header_alignment
            ws[f'F{row_num}'].border = border

            ws[f'G{row_num}'].value = value_added_sum
            ws[f'G{row_num}'].font = header_font
            ws[f'G{row_num}'].alignment = header_alignment
            ws[f'G{row_num}'].border = border

            ws[f'H{row_num}'].value = total_price_sum
            ws[f'H{row_num}'].font = header_font
            ws[f'H{row_num}'].alignment = header_alignment
            ws[f'H{row_num}'].border = border

            # Add footer with signatures
            footer_start_row = row_num + 2
            ws.merge_cells(f'A{footer_start_row}:K{footer_start_row + 2}')
            footer_cell = ws[f'A{footer_start_row}']
            footer_cell.value = 'تنخواه‌دار: ' + cover.user.get_full_name() + '\n' + \
                               'عامل مالی: فاطمه کرامتی\n' + \
                               cover.unit.title + ': ' + cover.unit.manager.get_full_name() + '\n' + \
                               'مدیرکل اداری مالی: بنیامین عبداللهی\n' + \
                               'معاون برنامه‌ریزی، پشتیبانی و توسعه منابع انسانی: مجید جاودانی'
            footer_cell.font = regular_font
            footer_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Adjust column widths
            column_widths = [8, 10, 20, 12, 15, 15, 15, 15, 10, 20, 30]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # Create HTTP response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="cover{cover.id}.xlsx"'
            wb.save(response)

            return response
        return HttpResponse(status=403)



class AddOrUpdateInvoice(GenericAPIView):
    permission_classes = [IsInvoiceCollector]

    def post(self, request):
        cover = get_object_or_404(InvoiceCover, id=request.data['cover'], user=request.user, locked=False)
        if request.data['id'] or request.user.groups.filter(name='invoice-confirm1').exists():
            invoice = get_object_or_404(Invoice, id=request.data['id'], cover_id=cover)
            invoice.unit_id = request.data['unit']
            invoice.no = request.data['no']
            invoice.date = request.data['date']
            invoice.price = request.data['price']
            invoice.category_id = request.data['category']
            invoice.has_paper = request.data['has_paper']
            invoice.value_added = request.data['value_added']
            invoice.issuer = request.data['issuer']
            invoice.description = request.data['description']
            invoice.save()
        else:
            invoice = Invoice.objects.create(cover=cover, unit_id=request.data['unit'], no=request.data['no'], date=request.data['date'], price=request.data['price'], category_id=request.data['category'], has_paper=request.data['has_paper'], value_added=request.data['value_added'], issuer=request.data['issuer'], description=request.data['description'])

        return Response(data=SerInvoice(invoice).data)


class RemoveInvoice(DestroyAPIView):
    permission_classes = [IsInvoiceCollector]

    def get_queryset(self):
        return Invoice.objects.filter(id=self.kwargs['pk'], cover__user=self.request.user, cover__locked=False)


class LockCover(GenericAPIView):
    permission_classes = [IsInvoiceCollector]

    def post(self, request):
        cover = get_object_or_404(InvoiceCover, id=request.data['id'], user=request.user)
        if cover.locked is False or cover.accepted is False:
            cover.locked = True
            cover.lock_time = jdatetime.datetime.now()
            _status = 'قفل روکش'
            if cover.accepted is False:
                cover.accepted = None
                _status = 'ارسال مجدد'
            if cover.confirm1 is False:
                cover.confirm1 = None
                _status = 'ارسال مجدد'
            if cover.confirm2 is False:
                cover.confirm2 = None
                _status = 'ارسال مجدد'
            if cover.confirm3 is False:
                cover.confirm3 = None
                _status = 'ارسال مجدد'
            cover.save()
            cover.tasks.create(user=request.user, status=_status)
            return Response(data={'cover': SerInvoiceCover(cover).data, 'todo_invoice': request.user.todo_invoice})
        return Response(data='شما دسترسی لازم ندارید', status=status.HTTP_403_FORBIDDEN)


class ToggleCoverAccepted(GenericAPIView):
    def post(self, request):
        cover = get_object_or_404(InvoiceCover, id=request.data['id'])
        if cover.locked is True and cover.confirm1 is None and request.user.groups.filter(name='invoice-confirm1').exists():
            cover.confirm1 = request.data['accept']
            cover.confirm1_note = request.data['note']
            cover.confirm1_time = jdatetime.datetime.now()
            _status = 'بررسی کارشناس مالی'
        elif cover.confirm1 is True and cover.confirm2 is None and request.user.is_head_of_unit:
            cover.confirm2 = request.data['accept']
            cover.confirm2_note = request.data['note']
            cover.confirm2_time = jdatetime.datetime.now()
            _status = 'بررسی مدیر واحد'
        elif cover.confirm2 is True and cover.confirm3 is None and request.user.groups.filter(name='invoice-confirm3').exists():
            cover.confirm3 = request.data['accept']
            cover.confirm3_note = request.data['note']
            cover.confirm3_time = jdatetime.datetime.now()
            _status = 'تأیید معاون برنامه‌ریزی'
        elif cover.confirm3 is True and cover.accepted is None and request.user.groups.filter(name='invoice-accept').exists():
            cover.accepted = request.data['accept']
            cover.accept_note = request.data['note']
            cover.accept_time = jdatetime.datetime.now()
            _status = 'دستور پرداخت'
        else:
            return Response(data='شما دسترسی لازم ندارید', status=status.HTTP_403_FORBIDDEN)
        if not request.data['accept']:
            cover.locked = False
        cover.save()
        cover.tasks.create(user=request.user, status=_status, accept=request.data['accept'], note=request.data['note'])
        return Response(data={'cover': SerInvoiceCover(cover).data, 'todo_invoice': request.user.todo_invoice})


class RemoveCover(DestroyAPIView):
    permission_classes = [IsInvoiceCollector]

    def get_queryset(self):
        return InvoiceCover.objects.filter(id=self.kwargs['pk'], user=self.request.user, locked=False)


class InvoiceCoverUserList(ListAPIView):
    permission_classes = [IsInvoiceCollector]
    serializer_class = SerUserList

    def get_queryset(self):
        return User.objects.annotate(val=Count('invoicecover', filter=Q(invoicecover__locked=True))).filter(val__gt=0).distinct()


class AddOrUpdateInvoiceCover(GenericAPIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if request.data['type'] == 'مستقیم' and not request.user.groups.filter(name='direct-invoice-registrar').exists():
            if not request.user.groups.filter(name='invoice-confirm1').exists():
                return Response(data='شما دسترسی ثبت روکش مستقیم ندارید', status=status.HTTP_400_BAD_REQUEST)
        if bool(int(request.data.get('id', 0))):
            if request.user.groups.filter(name='invoice-confirm1').exists():
                print("sssssssssssssssssssssssssssssss")
                cover = get_object_or_404(InvoiceCover, id=request.data['id'])
            else:
                cover = get_object_or_404(InvoiceCover, id=request.data['id'], user=request.user)
        else:
            cover = InvoiceCover.objects.create(user=request.user, unit=request.user.post.unit)
        if cover.locked and cover.confirm1 is not False and cover.confirm2 is not False and cover.confirm3 is not False and cover.accepted is not False:
            return Response(data='روکش موردنظر قابل ویرایش نیست', status=status.HTTP_400_BAD_REQUEST)
        cover.type = request.data['type']
        if request.data['type'] == 'مستقیم':
            cover.setadiran = request.data['setadiran'] == 'true'
            cover.sheba = request.data['sheba']
            cover.sheba_owner = request.data.get('sheba_owner', )
            if 'business_license' in request.data:
                cover.business_license = request.data['business_license']
            else:
                if not request.data['business_license_url']:
                    cover.business_license = None
            if 'id_card' in request.data:
                cover.id_card = request.data['id_card']
            else:
                if not request.data['id_card_url']:
                    cover.id_card = None
            if 'factor' in request.data:
                cover.factor = request.data['factor']
            else:
                if not request.data['factor_url']:
                    cover.factor = None
        else:
            cover.setadiran = False
            cover.sheba = None
            cover.sheba_owner = None
            cover.business_license = None
            cover.id_card = None
            cover.factor = None
        cover.save()
        return Response(data=SerInvoiceCover(cover).data)


class DepositDirectInvoiceCover(GenericAPIView):
    def post(self, request):
        if not request.user.groups.filter(name='invoice-deposit').exists():
            return Response(data='شما دسترسی لازم ندارید', status=status.HTTP_403_FORBIDDEN)
        cover = get_object_or_404(InvoiceCover, id=request.data['id'], accepted=True, type='مستقیم', deposit_time=None)
        cover.slip = request.data['slip']
        cover.deposit_time = jdatetime.datetime.now()
        cover.save()
        cover.tasks.create(user=request.user, status='پرداخت')
        return Response(data={'cover': SerInvoiceCover(cover).data, 'todo_invoice': request.user.todo_invoice})


class InvoiceCoverTaskList(ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = SerInvoiceCoverTask

    def get_queryset(self):
        user = self.request.user
        if user.groups.filter(name__in=['invoice-confirm1', 'invoice-confirm3', 'invoice-accept', 'invoice-deposit']).exists():
            cover = get_object_or_404(InvoiceCover, id=self.kwargs['pk'])
        elif user.is_head_of_unit:
            cover = get_object_or_404(InvoiceCover, id=self.kwargs['pk'], unit=user.post.unit)
        else:
            cover = get_object_or_404(InvoiceCover, id=self.kwargs['pk'], user=user)
        return cover.tasks.all()
