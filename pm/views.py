import operator
import datetime
import jdatetime
from io import BytesIO

from prj.models import Project
from .serializers import *
from weasyprint import HTML
from openpyxl import styles
from openpyxl import Workbook
from core.gsm import send_sms
from django.apps import apps
from functools import reduce
from core.permissions import *
from django.db.models import Q
from collections import Counter
from rest_framework import status
from rest_framework.views import APIView
from core.models import User, Notification
from rest_framework.response import Response
from django.http import FileResponse, HttpResponse
from django.template.loader import render_to_string
from rest_framework.permissions import IsAuthenticated
from openpyxl.worksheet.table import Table, TableStyleInfo
from django.contrib.auth.mixins import UserPassesTestMixin
from .models import Task, Job, JobAppendix, JobChat, Approval, Session, Tag, FlowPatternType
from rest_framework.generics import GenericAPIView, ListAPIView, get_object_or_404, RetrieveAPIView, UpdateAPIView, \
    DestroyAPIView, CreateAPIView

from django.db import DataError
from django.core.exceptions import ValidationError
from core.serializers import SerUser

def get_today(request):
    return HttpResponse(str(jdatetime.datetime.now().date()))


def is_senior_manager(manager_user, target_user):
    """
    بررسی می‌کند که آیا کاربر manager_user مدیر ارشد کاربر target_user است یا نه
    این تابع سلسله مراتب را تا 3 سطح بالا بررسی می‌کند
    """
    if not manager_user.post or not target_user.post:
        return False

    # چک می‌کنیم که آیا target_user در زیرمجموعه manager_user است
    # بررسی سطح 1: مدیر مستقیم
    if target_user.post.parent == manager_user.post:
        return True

    # بررسی سطح 2: مدیر غیرمستقیم (2 سطح بالا)
    if target_user.post.parent and target_user.post.parent.parent == manager_user.post:
        return True

    # بررسی سطح 3: مدیر غیرمستقیم (3 سطح بالا)
    if (target_user.post.parent and
        target_user.post.parent.parent and
        target_user.post.parent.parent.parent == manager_user.post):
        return True

    return False



def get_manager(user):
    post = user.post
    while post.parent:
        if post.parent.title == 'رئیس مرکز':
            return False
        post = post.parent

    manager = User.objects.get(post=post)
    return manager


def has_manager_access(creator, user):
    manager = get_manager(creator)
    if manager == user:
        return True
    return False

def is_manager_user(user):

    return user.post.is_manager


class MyFellowList(ListAPIView):
    serializer_class = SerMembers

    def get_queryset(self):
        user = self.request.user
        if self.request.user.username == 'javdani':  # جاودانی به همه نیروهای معاونت ارجاع دهد
            user_list = User.objects.filter(is_active=True, post__isnull=False).filter(
                post__unit_id__in=[7, 20, 21, 22, 28]).exclude(post__level__in=['خدمات', 'راننده'])
        else:
            user_list = User.objects.filter(is_active=True, post__isnull=False).filter(
                Q(post__parent=user.post) | Q(post=user.post) | Q(fellowed__fellower=self.request.user))
        return user_list.distinct()


def get_task_list_old(user, archive, team):
    tasks = SerTaskList(user.tasks.filter(job__archive=archive), many=True).data
    informees_tasks = SerTaskList(Task.objects.filter( job__informees=user.pk, job__archive=archive).order_by('job_id').distinct(
            'job_id'), many=True, ).data
    tasks += informees_tasks
    for task in tasks:
        #if user.post.is_manager or task['is_owner']:
        task['job']['is_owner'] = True
        #else:
        #    task['job']['is_owner'] = False


        if task in informees_tasks:
            task['job']['is_informees'] = True
        else:
            task['job']['is_informees'] = False


    data = [{
        'id': tag.id,
        'title': tag.title,
        'todo': [t for t in filter(lambda i: i['job']['status'] == 'todo' and i['tag'] == tag.id, tasks)],
        'doing': [t for t in filter(lambda i: i['job']['status'] == 'doing' and i['tag'] == tag.id, tasks)],
        'done': [t for t in filter(lambda i: i['job']['status'] == 'done' and i['tag'] == tag.id, tasks)],
    } for tag in user.tag_set.all()]
    data0 = {
        'id': None,
        'title': 'کارهای من',
        'todo': [t for t in filter(lambda i: i['job']['status'] == 'todo' and i['tag'] is None, tasks)],
        'doing': [t for t in filter(lambda i: i['job']['status'] == 'doing' and i['tag'] is None, tasks)],
        'done': [t for t in filter(lambda i: i['job']['status'] == 'done' and i['tag'] is None, tasks)]
    }
    if team:
        jobs = list(map(lambda i: i['job']['id'], tasks))
        task_list = Task.objects.filter(job__archive=archive).filter(
            Q(user__post__parent=user.post) | Q(user__post__parent__parent=user.post) | Q(
                user__post__parent__parent__parent=user.post)).exclude(job__in=jobs).order_by('job_id').distinct(
            'job_id')


        tasks = SerTaskList(task_list, many=True).data
        for task in tasks:
            task['job']['is_owner'] = True
        #tasks = list(map(lambda i: {**i, 'is_owner': False, 'is_seen': True}, tasks))
        data0['todo'].extend([t for t in filter(lambda i: i['job']['status'] == 'todo', tasks)], )
        data0['doing'].extend([t for t in filter(lambda i: i['job']['status'] == 'doing', tasks)], )
        data0['done'].extend([t for t in filter(lambda i: i['job']['status'] == 'done', tasks)], )
    data.insert(0, data0)
    return data


from django.db.models import Q

def get_task_list_v1(user, archive, team):

    # --- تسک‌های مستقیم کاربر ---
    user_tasks_qs = user.tasks.filter(job__archive=archive)
    tasks = SerTaskList(user_tasks_qs, many=True).data

    # --- تسک‌هایی که کاربر مطلع آن‌هاست ---
    informees_qs = Task.objects.filter(
        job__informees=user.pk,
        job__archive=archive
    ).order_by('job_id').distinct('job_id')

    informees_tasks = SerTaskList(informees_qs, many=True).data
    tasks += informees_tasks

    # --- تنظیم فلگ‌ها ---
    for task in tasks:
        task['job']['is_owner'] = True
        task['job']['is_informees'] = task in informees_tasks

    # --- گرفتن id تگ‌های متعلق به کاربر ---
    user_tag_ids = set(user.tag_set.values_list('id', flat=True))

    # --- ساخت لیست تگ‌ها ---
    data = [{
        'id': tag.id,
        'title': tag.title,
        'todo': [
            t for t in tasks
            if t['job']['status'] == 'todo' and t['tag'] == tag.id
        ],
        'doing': [
            t for t in tasks
            if t['job']['status'] == 'doing' and t['tag'] == tag.id
        ],
        'done': [
            t for t in tasks
            if t['job']['status'] == 'done' and t['tag'] == tag.id
        ],
    } for tag in user.tag_set.all()]

    # --- کارهای من (بدون تگ یا تگ غیر متعلق به کاربر) ---
    data0 = {
        'id': None,
        'title': 'کارهای من',
        'todo': [
            t for t in tasks
            if t['job']['status'] == 'todo' and
               (t['tag'] is None or t['tag'] not in user_tag_ids)
        ],
        'doing': [
            t for t in tasks
            if t['job']['status'] == 'doing' and
               (t['tag'] is None or t['tag'] not in user_tag_ids)
        ],
        'done': [
            t for t in tasks
            if t['job']['status'] == 'done' and
               (t['tag'] is None or t['tag'] not in user_tag_ids)
        ]
    }

    # --- اگر شامل تیم باشد ---
    if team:
        existing_job_ids = [t['job']['id'] for t in tasks]

        team_qs = Task.objects.filter(
            job__archive=archive
        ).filter(
            Q(user__post__parent=user.post) |
            Q(user__post__parent__parent=user.post) |
            Q(user__post__parent__parent__parent=user.post)
        ).exclude(
            job__in=existing_job_ids
        ).order_by('job_id').distinct('job_id')

        team_tasks = SerTaskList(team_qs, many=True).data

        for task in team_tasks:
            task['job']['is_owner'] = True
            task['job']['is_informees'] = False

        # فقط طبق همان منطق «کارهای من» اضافه شوند
        data0['todo'].extend([
            t for t in team_tasks
            if t['job']['status'] == 'todo' and
               (t['tag'] is None or t['tag'] not in user_tag_ids)
        ])

        data0['doing'].extend([
            t for t in team_tasks
            if t['job']['status'] == 'doing' and
               (t['tag'] is None or t['tag'] not in user_tag_ids)
        ])

        data0['done'].extend([
            t for t in team_tasks
            if t['job']['status'] == 'done' and
               (t['tag'] is None or t['tag'] not in user_tag_ids)
        ])

    # اضافه کردن کارهای من به ابتدای لیست
    data.insert(0, data0)

    return data

from django.db.models import Q

def get_task_list_v3(user, archive, team):

    # -------------------------
    # 1️⃣ job های مربوط به خود کاربر
    # -------------------------
    user_job_ids = Task.objects.filter(
        user=user,
        job__archive=archive
    ).values_list('job_id', flat=True)

    # -------------------------
    # 2️⃣ job هایی که کاربر مطلع است
    # -------------------------
    informee_job_ids = Job.objects.filter(
        informees=user,
        archive=archive
    ).values_list('id', flat=True)

    # -------------------------
    # 3️⃣ اگر شامل تیم باشد
    # -------------------------
    team_job_ids = []

    if team:
        team_users = User.objects.filter(
            Q(post__parent=user.post) |
            Q(post__parent__parent=user.post) |
            Q(post__parent__parent__parent=user.post)
        )

        team_job_ids = Task.objects.filter(
            user__in=team_users,
            job__archive=archive
        ).values_list('job_id', flat=True)

    # -------------------------
    # 4️⃣ ترکیب همه job ها
    # -------------------------
    all_job_ids = set(user_job_ids) | set(informee_job_ids) | set(team_job_ids)

    # گرفتن فقط یک تسک از هر job برای نمایش
    tasks_qs = Task.objects.filter(
        job_id__in=all_job_ids
    ).order_by('job_id').distinct('job_id')

    tasks = SerTaskList(tasks_qs, many=True).data

    # -------------------------
    # فلگ‌ها
    # -------------------------
    informee_job_ids_set = set(informee_job_ids)

    for task in tasks:
        task['job']['is_owner'] = True
        task['job']['is_informees'] = task['job']['id'] in informee_job_ids_set

    # -------------------------
    # تگ‌های کاربر
    # -------------------------
    user_tag_ids = set(user.tag_set.values_list('id', flat=True))

    # -------------------------
    # ستون تگ‌ها
    # -------------------------
    data = [{
        'id': tag.id,
        'title': tag.title,
        'todo': [t for t in tasks if t['job']['status'] == 'todo' and t['tag'] == tag.id],
        'doing': [t for t in tasks if t['job']['status'] == 'doing' and t['tag'] == tag.id],
        'done': [t for t in tasks if t['job']['status'] == 'done' and t['tag'] == tag.id],
    } for tag in user.tag_set.all()]

    # -------------------------
    # کارهای من
    # -------------------------
    data0 = {
        'id': None,
        'title': 'کارهای من',
        'todo': [
            t for t in tasks
            if t['job']['status'] == 'todo'
            and (t['tag'] is None or t['tag'] not in user_tag_ids)
        ],
        'doing': [
            t for t in tasks
            if t['job']['status'] == 'doing'
            and (t['tag'] is None or t['tag'] not in user_tag_ids)
        ],
        'done': [
            t for t in tasks
            if t['job']['status'] == 'done'
            and (t['tag'] is None or t['tag'] not in user_tag_ids)
        ],
    }

    data.insert(0, data0)

    return data

from django.db.models import Q


def get_all_subordinates(manager):
    """
    پیدا کردن تمام زیرمجموعه‌های مدیر به صورت recursive
    بدون محدودیت سطح
    """
    if not manager.post:
        return User.objects.none()

    subordinates = []
    stack = [manager.post]

    while stack:
        current_post = stack.pop()

        children_posts = Post.objects.filter(parent=current_post)
        for child in children_posts:
            stack.append(child)

        users = User.objects.filter(post=current_post)
        subordinates.extend(users)

    return User.objects.filter(id__in=[u.id for u in subordinates]).exclude(id=manager.id)


def get_task_list(user, archive, team):

    # ---------------------------
    # 1️⃣ Jobهای مستقیم کاربر
    # ---------------------------
    user_job_ids = set(
        Task.objects.filter(
            user=user,
            job__archive=archive
        ).values_list('job_id', flat=True)
    )

    # ---------------------------
    # 2️⃣ Jobهایی که کاربر مطلع است
    # ---------------------------
    informee_job_ids = set(
        Job.objects.filter(
            informees=user,
            archive=archive
        ).values_list('id', flat=True)
    )

    # ---------------------------
    # 3️⃣ Jobهای تیم
    # ---------------------------
    team_job_ids = set()

    if team:
        team_users = get_all_subordinates(user)

        if team_users.exists():
            team_job_ids = set(
                Task.objects.filter(
                    user__in=team_users,
                    job__archive=archive
                ).values_list('job_id', flat=True)
            )

    # ---------------------------
    # 4️⃣ ترکیب همه jobها
    # ---------------------------
    all_job_ids = user_job_ids | informee_job_ids | team_job_ids

    if not all_job_ids:
        return []

    tasks_qs = (
        Task.objects
        .filter(job_id__in=all_job_ids)
        .select_related('job', 'tag')
        .order_by('job_id')
        .distinct('job_id')   # یک تسک از هر job
    )

    tasks = SerTaskList(tasks_qs, many=True).data

    informee_job_ids_set = informee_job_ids
    user_tag_ids = set(user.tag_set.values_list('id', flat=True))

    # ---------------------------
    # فلگ‌ها
    # ---------------------------
    for task in tasks:
        task['job']['is_owner'] = True
        task['job']['is_informees'] = task['job']['id'] in informee_job_ids_set

    # ---------------------------
    # ستون تگ‌ها
    # ---------------------------
    data = []

    for tag in user.tag_set.all():
        data.append({
            'id': tag.id,
            'title': tag.title,
            'todo':  [t for t in tasks if t['job']['status'] == 'todo'  and t['tag'] == tag.id],
            'doing': [t for t in tasks if t['job']['status'] == 'doing' and t['tag'] == tag.id],
            'done':  [t for t in tasks if t['job']['status'] == 'done'  and t['tag'] == tag.id],
        })

    # ---------------------------
    # کارهای من
    # ---------------------------
    data0 = {
        'id': None,
        'title': 'کارهای من',
        'todo': [
            t for t in tasks
            if t['job']['status'] == 'todo'
            and (t['tag'] is None or t['tag'] not in user_tag_ids)
        ],
        'doing': [
            t for t in tasks
            if t['job']['status'] == 'doing'
            and (t['tag'] is None or t['tag'] not in user_tag_ids)
        ],
        'done': [
            t for t in tasks
            if t['job']['status'] == 'done'
            and (t['tag'] is None or t['tag'] not in user_tag_ids)
        ],
    }

    data.insert(0, data0)

    return data


class UpdateTagList(GenericAPIView):
    def post(self, request):
        try:
            ids = [t['id'] for t in request.data['tags']]
            request.user.tag_set.exclude(id__in=ids).delete()

            for data in request.data['tags']:
                title = data.get('title', '').strip()

                if not title:
                    return Response({'detail': 'فیلد title نباید خالی باشد.'}, status=status.HTTP_400_BAD_REQUEST)

                # ثبت تگ جدید
                if data['id'] == 0:
                    request.user.tag_set.create(title=title)
                # ویرایش تگ موجود
                else:
                    request.user.tag_set.filter(id=data['id']).update(title=title)

            return Response(
                data=get_task_list(request.user, request.data['archive'], request.data['team'],)
            )

        except (DataError, ValidationError):
            return Response(
                {'detail': 'طول عنوان بیش از حد مجاز است.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class TaskList(GenericAPIView):
    def get(self, request):
        archive = request.GET.get('archive', 'false') == 'true'
        team = request.GET.get('team', 'false') == 'true'
        return Response(data=get_task_list(request.user, archive, team))


class GetOrUpdateUserTagList(APIView):
    def get(self, request):
        user = request.user
        tags = Tag.objects.filter(user=user)
        data = SerTag(tags, many=True).data
        return Response(data=data)

    def post(self, request):
        user = request.user
        tags = Tag.objects.filter(user=user)
        tag_ids = request.data['tags']
        for order, tag_id in enumerate(tag_ids):
            tag = tags.get(id=tag_id)
            tag.order = order
            tag.save()

        tags = Tag.objects.filter(user=user)
        data = SerTag(tags, many=True).data
        return Response(data=data)


class UpdateTasksStatusAndOrder(GenericAPIView):
    def post(self, request):
        for tag in ['todo', 'doing', 'done']:
            for (order, pk) in enumerate(request.data[tag]):
                #task = get_object_or_404(Task, id=pk, user=request.user)
                task = Task.objects.filter(pk=pk, user=request.user).first()
                if request.user.post.is_manager and not task:
                    task = get_object_or_404(Task, pk=pk)

                if task:
                    task.job.status = tag
                    task.job.save()
                try:
                    task.order = order
                    task.save()
                except Exception as e:
                    pass
        return Response(data='done')


class TaskDetail(GenericAPIView):
    def get(self, request, pk):
        task = get_object_or_404(Task, id=pk)
        access_task = Task.objects.filter(job__informees=request.user.pk)
        try:
            if access_task or task.user == request.user or task.user.post.parent == request.user.post or task.user.post.parent.parent == request.user.post or task.user.post.parent.parent.parent == request.user.post:
                task.is_seen = True
                chat = task.job.chats.first()
                if chat:
                    task.last_seen_chat = chat
                task.save()
                request.user.notifications.filter(task=task).update(seen_time=jdatetime.datetime.now())
                data = {
                    'data': SerTaskDetail(task).data,
                    'todo_task': request.user.todo_task,
                }

                if task.user == request.user and task.is_owner:
                    data['data']['is_owner'] = True

                else:
                    data['data']['is_owner'] = request.user.post.is_manager

                return Response(data=data)
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(data='شما دسترسی لازم ندارید', status=status.HTTP_403_FORBIDDEN)


class RemoveJob(DestroyAPIView):
    def get_object(self):
        job = get_object_or_404(Job, pk=self.kwargs['pk'])
        if job.tasks.filter(is_owner=True, user=self.request.user).exists():
            return job
        return None


class UpdateJob(UpdateAPIView):
    serializer_class = SerJob

    def get_queryset(self):
        user = self.request.user
        # اگر فقط وضعیت تغییر می‌کند، هر کسی که در تسک است می‌تواند تغییر دهد
        #if 'status' in self.request.data:
        #    return Job.objects.filter(tasks__user=user)

        # برای ویرایش کامل، مالک یا مدیران ارشد می‌توانند ویرایش کنند
        # پیدا کردن تمام تسک‌هایی که کاربر مالک است یا مدیر ارشد مالک است
        owner_jobs = Job.objects.filter(tasks__user=user, tasks__is_owner=True)

        # اگر کاربر post ندارد، فقط تسک‌های مالک را برمی‌گردانیم
        #if not user.post:
        #    return owner_jobs

        # پیدا کردن تسک‌هایی که کاربر مدیر ارشد مالک آن‌ها است
        # باید تمام تسک‌هایی را پیدا کنیم که مالک آن‌ها در زیرمجموعه کاربر است
        from django.db.models import Q
        senior_manager_jobs = Job.objects.filter(
            Q(tasks__is_owner=True) &
            (
                Q(tasks__user__post__parent=user.post) |
                Q(tasks__user__post__parent__parent=user.post) |
                Q(tasks__user__post__parent__parent__parent=user.post)
            )
        )

        isj =  False
        if self.request.user.username == 'javdani':
            isj = True

        return Job.objects.all()

    def perform_update(self, serializer):
        instance = serializer.instance
        old_instance = Job.objects.get(pk=instance.pk)
        user = self.request.user

        # بررسی اینکه آیا کاربر مالک است یا مدیر ارشد
        is_owner = instance.tasks.filter(user=user, is_owner=True).exists()
        is_manager = is_manager_user(self.request.user)
        owner_task = None

        if not is_owner:
            # پیدا کردن مالک تسک
            owner_task = instance.tasks.filter(is_owner=True).first()

        # ذخیره تغییرات
        serializer.save()

        # اگر مدیر ارشد ویرایش کرده (نه مالک)، لاگ را ثبت می‌کنیم
        if is_manager and not is_owner and owner_task:
            changes = []

            # بررسی تغییرات در فیلدهای مختلف
            if old_instance.title != instance.title:
                changes.append(f'عنوان از «{old_instance.title}» به «{instance.title}» تغییر یافت')

            if old_instance.note != instance.note:
                changes.append('شرح تغییر یافت')

            if old_instance.deadline != instance.deadline:
                old_deadline = old_instance.deadline.strftime('%Y/%m/%d') if old_instance.deadline else 'ندارد'
                new_deadline = instance.deadline.strftime('%Y/%m/%d') if instance.deadline else 'ندارد'
                changes.append(f'مهلت از {old_deadline} به {new_deadline} تغییر یافت')

            if old_instance.status != instance.status:
                status_map = {'todo': 'انجام نشده', 'doing': 'در حال انجام', 'done': 'انجام شده'}
                old_status = status_map.get(old_instance.status, old_instance.status)
                new_status = status_map.get(instance.status, instance.status)
                changes.append(f'وضعیت از «{old_status}» به «{new_status}» تغییر یافت')

            if old_instance.urgency != instance.urgency:
                urgency_map = {1: 'عادی', 2: 'مهم', 3: 'خیلی مهم'}
                old_urgency = urgency_map.get(old_instance.urgency, old_instance.urgency)
                new_urgency = urgency_map.get(instance.urgency, instance.urgency)
                changes.append(f'فوریت از «{old_urgency}» به «{new_urgency}» تغییر یافت')

            if old_instance.suspended != instance.suspended:
                changes.append(f'وضعیت انتظار از «{"منتظر اقدام فرد دیگر" if old_instance.suspended else "عادی"}» به «{"منتظر اقدام فرد دیگر" if instance.suspended else "عادی"}» تغییر یافت')

            if old_instance.archive != instance.archive:
                changes.append(f'وضعیت بایگانی از «{"بایگانی شده" if old_instance.archive else "بایگانی نشده"}» به «{"بایگانی شده" if instance.archive else "بایگانی نشده"}» تغییر یافت')

            if old_instance.project_id != instance.project_id:
                changes.append('پروژه تغییر یافت')

            # اگر تغییری وجود داشت، پیام لاگ را ایجاد می‌کنیم
            if changes:
                owner_name = owner_task.user.get_full_name()
                body = f'ویرایش توسط مدیر ارشد ({user.get_full_name()}) - مالک تسک: {owner_name}\n'
                body += '\n'.join(f'• {change}' for change in changes)

                chat = instance.chats.create(user_id=0, body=body)

                # به‌روزرسانی last_seen_chat برای تمام تسک‌های مربوط به این کار
                for task in instance.tasks.all():
                    if task.user == user:
                        task.last_seen_chat = chat
                        task.save()

    def get_serializer_context(self, **kwargs):
        context = super().get_serializer_context()
        context['user'] = self.request.user
        return context

class UpdateJobDeadlineByMember(APIView):
    def post(self, request):
        task = get_object_or_404(Task, id=request.data['id'], user=request.user, is_owner=False)
        job = task.job
        if job.deadline:
            body = f'تغییر مهلت توسط {request.user.get_full_name()} از {job.deadline.strftime('%d-%m-%Y')} به {request.data['deadline'][8:]}-{request.data['deadline'][5:7]}-{request.data['deadline'][:4]}'
        else:
            body = f'تغییر مهلت توسط {request.user.get_full_name()} به {request.data['deadline'][8:]}-{request.data['deadline'][5:7]}-{request.data['deadline'][:4]}'
        chat = job.chats.create(user_id=0, body=body)
        job.deadline = request.data['deadline']
        job.save()
        task.last_seen_chat = chat
        task.save()
        return Response(data={'respite': job.respite, 'chat': SerJobChat(chat).data})


class AddJobAppendix(CreateAPIView, UserPassesTestMixin):
    serializer_class = SerJobAppendix
    queryset = JobAppendix.objects.all()

    def test_func(self):
        job = get_object_or_404(Job, id=self.request.data['job'])
        return job.tasks.filter(is_owner=True, user=self.request.user).exists()


class RemoveJobAppendix(DestroyAPIView):
    def get_object(self):
        appendix = get_object_or_404(JobAppendix, pk=self.kwargs['pk'])
        if appendix.job.tasks.filter(is_owner=True, user=self.request.user).exists():
            return appendix
        return None


class JobChatAdd(GenericAPIView):
    def post(self, request):
        job = get_object_or_404(Job, pk=request.data['job'])
        isj = False
        if self.request.user.username == 'javdani':
            isj = True

        access_task = Task.objects.filter(job__informees=request.user.pk, job__pk=job.id)
        if job.tasks.filter(user=request.user).exists() or job.tasks.filter(
                user__post__parent=request.user.post.parent).exists() or job.tasks.filter(
                user__post__parent__parent=request.user.post.parent).exists() or job.tasks.filter(
                user__post__parent__parent__parent=request.user.post.parent).exists() or access_task.exists() or isj:
            chat = job.chats.create(user=request.user)
            if 'body' in request.data:
                chat.body = request.data['body']
            if 'file' in request.data:
                chat.file = request.data['file']
            chat.save()
            for task in job.tasks.all():
                if task.user == request.user:
                    task.last_seen_chat = chat
                    task.save()
                else:
                    Notification.objects.create(user=task.user, title='پیام جدید در وظیفه', body=job.title,
                                                url=f'/task?id={task.id}', job_chat=chat, task=task)
            return Response(data=SerJobChat(chat).data)
        return Response(data='شما دسترسی لازم ندارید', status=status.HTTP_400_BAD_REQUEST)


class JobChatRemove(DestroyAPIView):

    def get_queryset(self):
        return self.request.user.jobchat_set.all()


class ChangeTaskTag(GenericAPIView):
    def post(self, request):
        task = get_object_or_404(Task, pk=request.data['task'], user=request.user)
        task.tag_id = request.data['tag']
        task.order = request.user.tasks.filter(job__archive=task.job.archive, tag=task.tag,
                                               job__status=task.job.status).count()
        task.save()
        return Response(data=SerTaskList(task).data)


class JobMembersChange(GenericAPIView):
    def post(self, request):

        #task = get_object_or_404(Task, id=request.data['task'], user=request.user)

        task = Task.objects.filter(pk=request.data['task'], user=request.user).first()
        job = None
        if task:
            job = task.job

        if request.user.post.is_manager and not job:
            task = get_object_or_404(Task, id=request.data['task'])
            job = task.job

        # حذف اعضای حذف شده
        job.tasks.exclude(user_id__in=request.data['users']).exclude(is_owner=True).exclude(is_owner=False,
                                                                                            user=request.user).delete()
        owner_task = job.tasks.filter(is_owner=True).first()
        # اصلاح عضویت مالک وظیفه در اعضای وظیفه
        if owner_task.user_id in request.data['users']:
            owner_task.is_committed = True
        else:
            owner_task.is_committed = False
        owner_task.save()
        # عضویت اعضای جدید
        added_ids = set(request.data['users']) - set(list(job.tasks.values_list('user_id', flat=True)))
        for pk in added_ids:
            user = User.objects.get(id=pk)

            # استفاده از تابع بهبود یافته برای پیدا کردن بهترین تگ مطابق
            member_tag = Tag.get_best_matching_tag(
                user=user,
                owner_tag_title=owner_task.tag.title if owner_task.tag else None,
                job_title=job.title
            )

            t = job.tasks.create(user_id=pk, tag=member_tag)
            Notification.objects.create(user_id=pk, title='وظیفه جدید', body=job.title, url=f'/task?id={t.id}', task=t)
        return Response(data=SerTaskListInJob(job.tasks, many=True).data)


class JobConfirmStatus(APIView):
    def get(self, request):
        job_id = request.GET.get('job_id')
        job = Job.objects.filter(id=job_id).first()
        if job:
            return Response({'confirm_status': job.confirm})
        return Response({'confirm_status': ""})


class JobAdd(GenericAPIView):
    def post(self, request):
        tag = request.data.get('tag', None)
        job = Job.objects.create(title=request.data['title'], status=request.data['status'],
                                 project_id=request.data['project'])
        order = request.user.tasks.filter(job__archive=False, tag=tag, job__status=status).count()
        task = job.tasks.create(user=request.user, tag_id=tag, order=order, is_owner=True,
                                is_committed=False if request.user.post else True)  # کاربران استانی میتوانند برای خودشان وظیفه تعریف کنند

        # دریافت تگ مالک تسک برای استفاده در انتخاب تگ مناسب برای اعضای دیگر
        owner_tag = None
        if tag:
            try:
                owner_tag = Tag.objects.get(id=tag, user=request.user)
            except Tag.DoesNotExist:
                pass

        for member in request.data['members']:
            if task.user.id == int(member):
                task.is_committed = True
                task.save()
            else:
                member_user = User.objects.get(id=member)
                # استفاده از تابع بهبود یافته برای پیدا کردن بهترین تگ مطابق
                member_tag = Tag.get_best_matching_tag(
                    user=member_user,
                    owner_tag_title=owner_tag.title if owner_tag else None,
                    job_title=job.title
                )
                job.tasks.create(user_id=member, tag=member_tag)
        return Response(data=SerTaskList(task).data)


class JobMedia(GenericAPIView):
    def post(self, request):
        isj = False
        if self.request.user.username == 'javdani':
            isj = True
        if request.data['field'] == 'appendix':

            appendix = get_object_or_404(JobAppendix, pk=request.data['pk'])
            access_task = Task.objects.filter(job__informees=request.user.pk, job__pk=appendix.job.id)
            if appendix.job.tasks.filter(Q(user=request.user) | Q(user__post__parent=request.user.post) | Q(
                    user__post__parent__parent=request.user.post) | Q(
                    user__post__parent__parent__parent=request.user.post)).exists() or access_task.exists() or isj:
                return FileResponse(appendix.file)
        if request.data['field'] == 'chat':

            chat = get_object_or_404(JobChat, pk=request.data['pk'])
            access_task = Task.objects.filter(job__informees=request.user.pk, job__pk=chat.job.id)
            if chat.job.tasks.filter(user=request.user).exists() or access_task.exists() or isj:
                return FileResponse(chat.file)
        return Response(data='شما دسترسی لازم ندارید', status=status.HTTP_403_FORBIDDEN)


class JobInformees(APIView):
    def get(self, request):
        qs = User.objects.filter(post__isnull=False)
        data = SerMembers(qs, many=True).data
        return Response(data=data)

    def post(self, request):
        data = request.data
        task_id = data.get('task')
        informees = data['informees']
        job = Task.objects.get(id=task_id).job
        job.informees.set([])
        if informees:
            job.informees.set(informees)
            job.save()

        try:
            data = []
            for us in job.informees.all():
                data.append({
                    'id': us.id,
                    'name': us.name,
                    'photo_url': us.photo_url,
                })

        except Exception as e:
            pass
        return Response(data=data)


class MySessionMemberList(ListAPIView):
    serializer_class = SerMembers

    def get_queryset(self):
        post = self.request.user.post
        if post.is_deputy:
            return User.objects.filter(
                Q(post__unit=post.unit) | Q(post__unit__parent=post.unit) | Q(post__is_manager=True))
        elif post.is_manager:
            return User.objects.filter(
                Q(post__unit=post.unit) | Q(post__unit__parent=post.unit.parent, post__is_manager=True) | Q(
                    post__unit=post.unit.parent, post__is_manager=True))
        elif post.id == 115:
            # مسئول دفتر معاونت توسعه برای مدیران معاونت بتواند جلسه تنظیم کند
            return User.objects.filter(post_id__in=[8, 21, 30, 32, 112, 240])
        else:
            return User.objects.filter(post__unit=post.unit)

class SessionListInMonth(GenericAPIView):
    def get(self, request):
        today = jdatetime.datetime.now().date()
        year = int(request.GET.get('year', 0)) or today.year
        month = int(request.GET.get('month', 0)) or today.month

        user = request.user

        base_qs = (
            Session.objects
            .select_related('room', 'user', 'unit', 'user__post', 'user__post__unit')
            .prefetch_related('members')
        )

        date_filters = {
            'date__gte': '{}-{}-01'.format(year, month),
            'date__lt': '{}-{}-01'.format(
                year + 1 if month == 12 else year,
                1 if month == 12 else month + 1
            )
        }
        base_qs = base_qs.filter(**date_filters)

        # برای استفاده دوباره: فیلتر «کاملاً تأیید شده»
        fully_approved_filter = (

            (Q(need_manager=False) | Q(manager_accept='تأیید')) &
            (Q(need_deputy=False) | Q(deputy_accept='تأیید'))
        )

        #  نقش‌ها
        if user.groups.filter(name='approval').exists():
            session_list = base_qs

        elif user.groups.filter(name='room_admin').exists():
            # ادمین اتاق:
            #  - تمام جلسات عمومی
            #  - همه جلساتی که خودش ایجاد کرده یا عضوش است
            session_list = base_qs.filter(
                Q(room__public=True) |
                Q(user=user) |
                Q(members__in=[user])
            )

        elif user.groups.filter(name='room_supervisor').exists():
            session_list = base_qs.filter(
                Q(members__in=[user]) | Q(user=user) | Q(room_agents__in=[user])
                | Q(photography_agents__in=[user]) | Q(filming_agents__in=[user])
                | Q(recording_agents__in=[user]) | Q(news_agents__in=[user])
                | Q(presentation_agents__in=[user])
            ).filter(fully_approved_filter)

        elif user.groups.filter(name='room_catering').exists():
            # قبلاً اینجا فقط fully_approved_filter AND ... بود
            # الان: جلسات زیر را نشان بده:
            #   1) جلسات کاملاً تأیید شده‌ای که برای پذیرایی مهم‌اند
            #   2) هر جلسه‌ای با اتاق عمومی که خودِ کاربر ثبت کرده (حتی اگر هنوز تأیید نشده)
            visibility_filter = (
                Q(members__in=[user]) |
                Q(user=user) |
                Q(room__public=True) |
                Q(breakfast__isnull=False) |
                Q(catering__isnull=False)
            )

            session_list = base_qs.filter(
                (fully_approved_filter & visibility_filter) |
                Q(user=user, room__public=True)   # این خط باگ رو برطرف می‌کند
            )

        elif user.post.is_deputy:
            session_list = base_qs.filter(
                Q(user__post__unit=user.post.unit) |
                Q(user__post__unit__parent=user.post.unit)
            )

        elif user.post.is_manager:
            session_list = base_qs.filter(user__post__unit=user.post.unit)

        else:
            session_list = base_qs.filter(
                Q(members__in=[user]) | Q(user=user) | Q(room_agents__in=[user]) |
                Q(photography_agents__in=[user]) | Q(filming_agents__in=[user]) |
                Q(recording_agents__in=[user]) | Q(news_agents__in=[user]) |
                Q(presentation_agents__in=[user])
            )

        session_list = session_list.distinct().order_by('room', 'date', 'start')

        date = jdatetime.date(year=year, month=month, day=1)
        _next = date + jdatetime.timedelta(days=32)
        prev = date - jdatetime.timedelta(days=5)

        data = {
            'year': year,
            'month': month,
            'title': '{} {}'.format(today.j_months_fa[date.month - 1], today.year),
            'next': {
                'year': _next.year,
                'month': _next.month,
                'title': '{} {}'.format(_next.j_months_fa[_next.month - 1], _next.year)
            },
            'prev': {
                'year': prev.year,
                'month': prev.month,
                'title': '{} {}'.format(prev.j_months_fa[prev.month - 1], prev.year)
            },
            'sessions': SerSessionList(
                instance=session_list,
                many=True,
                context={'user': user.id}
            ).data,
            'days': []
        }

        while date.month == month:
            data['days'].append(
                {'date': str(date), 'day': date.day, 'weekday': date.weekday(), 'week': date.weeknumber()}
            )
            date += jdatetime.timedelta(days=1)

        return Response(data=data)


class SessionList(GenericAPIView):
    """
    لیست جلسات بر اساس نقش کاربر، فیلترهای تاریخ، نوع جلسه، اتاق، دپارتمان و سرچ.
    - need_action = همه / منتظر اقدام
    """

    def get(self, request):
        user = request.user
        need_action = request.GET.get('need_action', 'همه')

        # کوئری پایه با related/prefetch برای کاهش تعداد کوئری‌ها
        base_qs = (
            Session.objects
            .select_related('room', 'user', 'unit', 'user__post', 'user__post__unit')
            .prefetch_related('members')
        )

        # تعریف شرایط "کامل تأیید شده" بودن جلسه
        fully_approved_filter = (
            (Q(room=None) | Q(accept_room=True)) &
            (Q(need_manager=False) | Q(manager_accept='تأیید')) &
            (Q(need_deputy=False) | Q(deputy_accept='تأیید'))
        )

        # اگر کاربر صفحه "منتظر اقدام" را می‌خواهد، صرفاً از todo_session_list استفاده می‌کنیم
        if need_action == 'منتظر اقدام':
            session_list = user.todo_session_list()

        else:
            # بر اساس نقش کاربر
            if user.groups.filter(name='approval').exists():
                # دسترسی ادمین ارشد: همه جلسات
                session_list = base_qs

            elif user.groups.filter(name='room_admin').exists():
                # ادمین سالن:
                # 1) همه جلسات سالن‌های عمومی (public room) که بعد از طی کامل زنجیره تأیید، نهایی شده‌اند
                # 2) جلساتی که خودش توش نقش دارد (مالک، عضو، ایجنت‌ها)
                # 3) درخواست‌های سالن عمومی که هنوز در انتظار تأیید هستند (accept_room=None)

                visibility = (
                    Q(members__in=[user]) |
                    Q(user=user) |
                    Q(room_agents__in=[user]) |
                    Q(photography_agents__in=[user]) |
                    Q(filming_agents__in=[user]) |
                    Q(recording_agents__in=[user]) |
                    Q(news_agents__in=[user]) |
                    Q(presentation_agents__in=[user]) |
                    Q(room__public=True)  # برنامه کامل سالن‌های عمومی
                )

                session_list = (
                    base_qs
                    .filter(
                        # جلسات تأییدشده که به کاربر ربط دارند
                        (fully_approved_filter & visibility)
                        |
                        # درخواست‌های رزرو سالن عمومی که هنوز در حال بررسی هستند
                        Q(room__public=True, accept_room=None)
                    )
                )

            elif user.groups.filter(name='room_supervisor').exists():
                # ناظر سالن: فقط جلسات کاملاً تأیید شده که به او ربط دارند
                session_list = (
                    base_qs
                    .filter(fully_approved_filter)
                    .filter(
                        Q(members__in=[user]) |
                        Q(user=user) |
                        Q(room_agents__in=[user]) |
                        Q(photography_agents__in=[user]) |
                        Q(filming_agents__in=[user]) |
                        Q(recording_agents__in=[user]) |
                        Q(news_agents__in=[user]) |
                        Q(presentation_agents__in=[user])
                    )
                )

            elif user.groups.filter(name='room_catering').exists():
                # ادمین پذیرایی:
                # 1) جلساتی که کامل تأیید شده‌اند و به او ربط دارند یا اتاق عمومی دارند
                # 2) جلسه‌های اتاق عمومی که خود کاربر ساخته حتی قبل از تأیید
                visibility = (
                    Q(members__in=[user]) |
                    Q(user=user) |
                    Q(room__public=True) |
                    Q(breakfast__isnull=False) |
                    Q(catering__isnull=False)
                )

                session_list = (
                    base_qs
                    .filter(
                        (fully_approved_filter & visibility)
                        |
                        Q(user=user, room__public=True)
                    )
                )

            elif getattr(user.post, 'is_deputy', False):
                # معاون: جلسات واحد خودش و واحدهای فرزند
                session_list = base_qs.filter(
                    Q(user__post__unit=user.post.unit) |
                    Q(user__post__unit__parent=user.post.unit)
                )

            elif getattr(user.post, 'is_manager', False):
                # مدیر: جلسات واحد خودش
                session_list = base_qs.filter(user__post__unit=user.post.unit)

            else:
                # کاربر معمولی: جلساتی که به نحوی در آن دخیل است
                session_list = base_qs.filter(
                    Q(members__in=[user]) |
                    Q(user=user) |
                    Q(room_agents__in=[user]) |
                    Q(photography_agents__in=[user]) |
                    Q(filming_agents__in=[user]) |
                    Q(recording_agents__in=[user]) |
                    Q(news_agents__in=[user]) |
                    Q(presentation_agents__in=[user])
                )

        # ---------------- فیلتر تاریخ ----------------
        from_date_str = request.GET.get('from_date', '')
        to_date_str = request.GET.get('to_date', '')
        today = jdatetime.date.today()

        if not from_date_str and not to_date_str:
            # اگر تاریخ ارسال نشده بود، فقط جلسات امروز به بعد
            session_list = session_list.filter(date__gte=today)
        else:
            if from_date_str:
                try:
                    from_date = jdatetime.datetime.strptime(from_date_str, '%Y-%m-%d').date()
                    session_list = session_list.filter(date__gte=from_date)
                except ValueError:
                    # اگر فرمت تاریخ اشتباه بود، نادیده بگیر
                    pass
            if to_date_str:
                try:
                    to_date = jdatetime.datetime.strptime(to_date_str, '%Y-%m-%d').date()
                    session_list = session_list.filter(date__lte=to_date)
                except ValueError:
                    pass

        # ---------------- فیلتر نوع جلسه ----------------
        session_type = request.GET.get('type', '')
        if session_type:
            try:
                session_list = session_list.filter(type=int(session_type))
            except ValueError:
                pass

        # ---------------- فیلتر اتاق ----------------
        room_filter = request.GET.get('room', '')
        if room_filter == 'with_room':
            session_list = session_list.filter(room__isnull=False)
        elif room_filter == 'without_room':
            session_list = session_list.filter(room__isnull=True)

        # ---------------- فیلتر دپارتمان و سرچ ----------------
        department = int(request.GET.get('department', 0) or 0)
        q = request.GET.get('q', None)

        if department:
            session_list = session_list.filter(
                Q(user__post__unit_id=department) |
                Q(user__post__unit__parent_id=department)
            )

        if q:
            q_list = q.split(' ')
            # سرچ بر اساس همه کلمات در عنوان
            q_filter = reduce(
                operator.and_,
                (Q(title__icontains=q_text) for q_text in q_list if q_text.strip())
            )
            session_list = session_list.filter(q_filter)

        # حذف رکوردهای تکراری و مرتب‌سازی
        session_list = session_list.distinct().order_by('date')

        # ---------------- صفحه‌بندی ----------------
        size = getattr(user.profile, 'page_size', 20)  # اگر page_size نداشت، دیفالت 20
        count = session_list.count()
        max_page = (count // size + 1) if count > 0 else 1

        try:
            page = int(request.GET.get('page', 1))
        except ValueError:
            page = 1

        page = max(1, min(page, max_page))

        start = size * (page - 1)
        end = size * page

        # ---------------- خروجی ----------------
        data = {
            'count': count,
            'page': page,
            'size': size,
            'q': q,
            '_q': q,
            'department': department,
            'need_action': need_action,
            # تعداد درخواست‌های سالن عمومی که هنوز تأیید نشده‌اند (برای نشان‌دادن نوتیف/بج)
            'pending_room_count': Session.objects.filter(room__public=True, accept_room=None).count() if user.groups.filter(name='room_admin').count()  else 0,
            'list': SerSessionList(
                instance=session_list[start:end],
                many=True,
                context={'user': user}
            ).data
        }

        return Response(data=data)
class SessionDetail(GenericAPIView):
    def get(self, request, pk):
        user = request.user
        session = get_object_or_404(Session, pk=pk)

        # Check if user has access to this session based on the same criteria used in the list views
        # Build the same query that would return this session in the list views
        session_qs = Session.objects.filter(pk=session.pk)  # Start with the specific session

        has_access = False

        if user.groups.filter(name='approval').exists():
            has_access = True
        elif user.groups.filter(name='room_admin').exists():
            has_access = session_qs.filter(
                Q(members__in=[user]) | Q(user=user) | Q(room__public=True)).exists()
        elif user.groups.filter(name='room_supervisor').exists():
            has_access = session_qs.filter(
                Q(members__in=[user]) | Q(user=user) | Q(room_agents__in=[user])
                | Q(photography_agents__in=[user]) | Q(filming_agents__in=[user])
                | Q(recording_agents__in=[user]) | Q(news_agents__in=[user])
                | Q(presentation_agents__in=[user])).filter(
                (Q(room=None) | Q(accept_room=True)) &
                (Q(need_manager=False) | Q(manager_accept='تأیید')) &
                (Q(need_deputy=False) | Q(deputy_accept='تأیید'))
            ).exists()
        elif user.groups.filter(name='room_catering').exists():
            has_access = session_qs.filter(
                (Q(room=None) | Q(accept_room=True)) &
                (Q(need_manager=False) | Q(manager_accept='تأیید')) &
                (Q(need_deputy=False) | Q(deputy_accept='تأیید')) &
                (Q(members__in=[user]) | Q(user=user) | Q(room__public=True) | Q(
                    breakfast__isnull=False) | Q(catering__isnull=False))
            ).exists()
        elif user.post.is_deputy:
            has_access = session_qs.filter(
                Q(user__post__unit=user.post.unit) | Q(user__post__unit__parent=user.post.unit)).exists()
        elif user.post.is_manager:
            has_access = session_qs.filter(user__post__unit=user.post.unit).exists()
        else:
            # For regular users, check if session is one they should have access to
            has_access = session_qs.filter(
                Q(members__in=[user]) | Q(user=user) | Q(room_agents__in=[user]) | Q(
                    photography_agents__in=[user]) | Q(filming_agents__in=[user]) | Q(
                    recording_agents__in=[user]) | Q(news_agents__in=[user]) | Q(
                    presentation_agents__in=[user])
            ).exists()

        if not has_access:
            if not user.groups.filter(name='room_catering').exists():
                return Response(data='شما دسترسی لازم ندارید', status=status.HTTP_403_FORBIDDEN)

        data = SerSessionDetail(session, context={'user': user}).data

        # Special handling for room_catering users - they should not see approvals
        if user.groups.filter(name='room_catering').exists()  and session.user != request.user:
            data['approvals'] = []

        # Add session creator information
        data['creator_info'] = {
            'id': session.user.id,
            'full_name': session.user.get_full_name(),
            'personnel_code': session.user.personnel_code,
            'post_title': session.user.post.title if session.user.post else '',
            'unit_title': session.user.post.unit.title if session.user.post and session.user.post.unit else '',
            'create_time': session.create_time.strftime('%Y-%m-%d %H:%M:%S'),
        }

        # Add catering details
        data['catering_details'] = {
            'need_breakfast': session.need_breakfast,
            'need_lunch': session.need_lunch,
            'need_catering': session.need_catering,
            'breakfast': session.breakfast,
            'lunch': session.lunch,
            'catering': session.catering,
            'attendee_count': session.attendee_count,
            'manager_accept': session.manager_accept,
            'manager_note': session.manager_note,
            'deputy_accept': session.deputy_accept,
            'deputy_note': session.deputy_note,
            'request_time': str(session.request_time) if session.request_time else None,
            'order_time': str(session.order_time) if session.order_time else None,
        }
        catering = request.user.groups.filter(name='room_catering').exists()
        data['can'] = {
            'edit': session.user == user or user in session.secretaries.all(),
            'secretaries': user in session.secretaries.all(),
            'catering': catering, #(session.room is None or session.accept_room is True) and (session.need_breakfast or session.need_lunch or session.need_catering) and not session.order_time and (not session.need_manager or session.manager_accept == 'تأیید') and (not session.need_deputy or session.deputy_accept == 'تأیید'),
            'manager_accept': session.need_manager and session.manager_accept == 'نامشخص' and user.is_head_of_unit,
            'deputy_accept': session.need_deputy and session.deputy_accept == 'نامشخص' and user.post.is_deputy,
            'rate': True if session.need_catering or session.need_breakfast or session.need_lunch else False,
        }
        return Response(data=data)


class SessionAddOrUpdate(GenericAPIView):
    def post(self, request):
        pk = int(request.data.get('id', 0))
        # خواندن ورودی‌های زمان/اتاق برای بررسی تداخل قبل از ایجاد یا بروزرسانی
        date = request.data.get('date')
        room_id = request.data.get('room')
        start_raw = request.data.get('start')
        end_raw = request.data.get('end')
        is_room_admin = request.user.groups.filter(name='room_admin').exists()
        accept_room_input = request.data.get('accept_room') if 'accept_room' in request.data else None
        manager_accept_input = request.data.get('manager_accept') if 'manager_accept' in request.data else None
        deputy_accept_input = request.data.get('deputy_accept') if 'deputy_accept' in request.data else None

        if room_id:
            try:
                # اعتبارسنجی پایه‌ای بازه زمانی
                if start_raw >= end_raw:
                    return Response(data={'detail': 'بازه زمانی نامعتبر است', 'code': 'invalid_time_range'},
                                    status=status.HTTP_400_BAD_REQUEST)

                room = get_object_or_404(Room, id=room_id)
                # برای اتاق‌های عمومی ۳۰ دقیقه حاشیه اضافه می‌کنیم
                if room.public:
                    start_for_check = (datetime.datetime.strptime(start_raw, '%H:%M') - datetime.timedelta(
                        minutes=30)).strftime('%H:%M')
                    end_for_check = (
                            datetime.datetime.strptime(end_raw, '%H:%M') + datetime.timedelta(minutes=30)).strftime(
                        '%H:%M')
                else:
                    start_for_check = start_raw
                    end_for_check = end_raw

                conflicts_qs = Session.objects.filter(room=room, date=date)
                if pk:
                    conflicts_qs = conflicts_qs.exclude(id=pk)
                conflicts_qs = conflicts_qs.exclude(accept_room=False)
                conflicts_qs = conflicts_qs.filter(end__gt=start_for_check, start__lt=end_for_check)
                if conflicts_qs.exists():
                    conflicts = [{
                        'title': s.title if (s.user == request.user or request.user in s.members.all()) else '-----',
                        'unit': s.user.post.unit.title,
                        'start': str(s.start)[:5],
                        'end': str(s.end)[:5]
                    } for s in conflicts_qs]
                    return Response(
                        data={'detail': 'تداخل زمانی با جلسات دیگر وجود دارد', 'code': 'overlap', 'list': conflicts},
                        status=status.HTTP_400_BAD_REQUEST)
            except Exception:
                return Response(data={'detail': 'ورودی زمان نامعتبر است', 'code': 'invalid_time_input'},
                                status=status.HTTP_400_BAD_REQUEST)

        # ادامه منطق اصلی — ایجاد یا بروزرسانی جلسه (بعد از بررسی تداخل)
        if pk:
            session = get_object_or_404(Session, id=pk, user=request.user)
            session.title = request.data['title']
            session.type = request.data['type']
            session.date = request.data['date']
            session.save()
        else:
            session = Session.objects.create(title=request.data['title'], type=request.data['type'],
                                             date=request.data['date'], user=request.user, unit=request.user.post.unit)
        session.refresh_from_db()
        session.week = session.date.weeknumber()
        session.sms = request.data['sms']
        print(request.data['members'])
        session.members.set(request.data['members'])
        session.secretaries.set(request.data.get('secretaries', []))
        session.agenda = request.data['agenda']

        # اگر پارامترهای اتاق/زمان تغییر کرده بود، برخی وضعیت‌ها ریست شوند
        if session.room_id != request.data['room'] or str(session.start) != str(request.data['start'] + ':00') or str(
                session.end) != str(request.data['end'] + ':00'):
            session.accept_room = None
            session.accept_photography = None
            session.accept_filming = None
            session.accept_recording = None
            session.accept_news = None
            session.accept_presentation = None
            session.room_agents.set([])
            session.photography_agents.set([])
            session.filming_agents.set([])
            session.recording_agents.set([])
            session.news_agents.set([])
            session.presentation_agents.set([])
            session.jobs.all().delete()
        if session.need_photography != request.data['need_photography']:
            session.accept_room = None
            session.accept_photography = None
            session.photography_agents.set([])
            session.jobs.filter(title='عکسبرداری جلسه').delete()
        if session.need_filming != request.data['need_filming']:
            session.accept_room = None
            session.accept_filming = None
            session.filming_agents.set([])
            session.jobs.filter(title='تصویربرداری جلسه').delete()
        if session.need_recording != request.data['need_recording']:
            session.accept_room = None
            session.accept_recording = None
            session.recording_agents.set([])
            session.jobs.filter(title='ضبط جلسه').delete()
        if session.need_news != request.data['need_news']:
            session.accept_room = None
            session.accept_news = None
            session.news_agents.set([])
            session.jobs.filter(title='تهیه خبر جلسه').delete()
        if session.need_presentation != request.data['need_presentation']:
            session.accept_room = None
            session.accept_presentation = None
            session.presentation_agents.set([])
            session.jobs.filter(title='هماهنگی ارائه جلسه').delete()
        session.room_id = request.data['room']
        session.place = request.data['place']

        # تنظیمات بر اساس عمومی/غیرعمومی بودن اتاق
        if session.room and session.room.public:
            session.need_photography = request.data['need_photography']
            session.need_filming = request.data['need_filming']
            session.need_recording = request.data['need_recording']
            session.need_news = request.data['need_news']
            session.need_presentation = request.data['need_presentation']
        else:
            session.need_photography = False
            session.need_filming = False
            session.need_recording = False
            session.need_news = False
            session.need_presentation = False

        # اگر جلسه در فردا یا بیشتر بود و جزئیات پذیرایی (تعداد کل اعضا یا اضافه شدن درخواست صبحانه یا ناهار یا پذیرایی) تغییر افزایشی داشت، فیلدهای فرآیند پذیرایی ریست میشود
        if session.date > jdatetime.date.today() and ((session.guest_count + session.members.count()) < (
                len(request.data['members']) + int(request.data['guest_count'])) or (
                                                              not session.need_breakfast and request.data[
                                                          'need_breakfast']) or (
                                                              not session.need_lunch and request.data[
                                                          'need_lunch']) or (not session.need_catering and request.data[
            'need_catering'])):
            session.order_time = None
            session.need_manager = False
            session.need_deputy = False
        session.start = request.data['start'] + ':00'
        session.end = request.data['end'] + ':00'
        session.members.set(request.data['members'])
        session.guest_count = request.data['guest_count']
        if session.date > jdatetime.date.today():
            session.need_breakfast = request.data['need_breakfast']
            session.need_lunch = request.data['need_lunch']
            session.need_catering = request.data['need_catering']
        if request.data['reset_catering_order'] and session.date > jdatetime.date.today():
            session.request_time = jdatetime.datetime.now()
            session.breakfast = None
            session.nf = None
            session.catering = []
            session.order_time = None
        if session.request_time is None:
            session.request_time = jdatetime.datetime.now()
        if session.manager_accept == 'عودت جهت اصلاح':
            session.manager_accept = 'نامشخص'
        if session.deputy_accept == 'عودت جهت اصلاح':
            session.deputy_accept = 'نامشخص'
        if accept_room_input is not None and is_room_admin:
            session.accept_room = accept_room_input
        if manager_accept_input is not None and session.need_manager and session.manager_accept == 'نامشخص' \
                and request.user.is_head_of_unit and session.is_room_approved:
            session.manager_accept = manager_accept_input
        if deputy_accept_input is not None and session.need_deputy and session.deputy_accept == 'نامشخص' \
                and request.user.post.is_deputy and session.is_room_approved and \
                (not session.need_manager or session.manager_accept == 'تأیید'):
            session.deputy_accept = deputy_accept_input
        session.save()

        # تنظیم برنامه‌ها (projects)
        if 'project' in request.data:
            project = Project.objects.filter(id=request.data['project']).first()
            session.project = project
            session.save()

        # مصوبات جلسه
        ids = list(map(lambda p: p['id'], request.data['approvals']))
        session.approvals.exclude(id__in=ids).delete()
        for item in request.data['approvals']:
            if item['id']:
                approval = get_object_or_404(Approval, id=item['id'])
                approval.title = item['title']
                approval.deadline = item['deadline'] if item['deadline'] else None
                approval.save()
                approval.members.set(item['members'])
            else:
                approval = Approval.objects.create(session=session, title=item['title'],
                                                   deadline=item['deadline'] if item['deadline'] else None)
                approval.members.set(item['members'])
        # ایجاد وظیفه برای هر مصوبه
        for approval in session.approvals.all():
            job = Job.objects.filter(approval=approval).first()
            if job is None:
                job = Job.objects.create(approval=approval, title=approval.title)
            job.deadline = approval.deadline
            job.title = approval.title
            job.save()
            if not job.tasks.filter(user=request.user).exists():
                job.tasks.create(user=request.user, is_owner=True, is_committed=request.user in approval.members.all())
            for member in approval.members.all():
                if not job.tasks.filter(user=member).exists():
                    job.tasks.create(user=member)
        return Response(data=SerSessionList(session).data)


class RemoveSession(DestroyAPIView):
    def get_queryset(self):
        return Session.objects.filter(user=self.request.user)


class SessionCateringAccept(APIView):
    def post(self, request):
        user = request.user
        session = get_object_or_404(Session, id=request.data['id'])
        if not session.is_room_approved:
            return Response(data='ابتدا رزرو اتاق باید تأیید شود', status=status.HTTP_400_BAD_REQUEST)
        if (session.need_manager and session.manager_accept == 'نامشخص' and user.is_head_of_unit
                and session.is_room_approved):
            session.manager_accept = request.data['manager_accept']
            session.manager_note = request.data['manager_note']
            session.save()
            return Response(data=SerSessionList(session).data)
        if (session.need_deputy and session.deputy_accept == 'نامشخص' and user.post.is_deputy
                and session.is_room_approved and (not session.need_manager or session.manager_accept == 'تأیید')):
            session.deputy_accept = request.data['deputy_accept']
            session.deputy_note = request.data['deputy_note']
            session.save()
            return Response(data=SerSessionList(session).data)
        return Response(data='شما دسترسی لازم ندارید', status=status.HTTP_403_FORBIDDEN)


class ApprovalList(GenericAPIView):
    def get(self, request):
        if request.user.groups.filter(name='approval').exists():
            session_list = Session.objects.all()
        elif request.user.post.is_deputy:
            session_list = Session.objects.filter(
                Q(user__post__unit=request.user.post.unit) | Q(user__post__unit__parent=request.user.post.unit) | Q(
                    members__in=[request.user])).distinct()
        elif request.user.post.is_manager:
            session_list = Session.objects.filter(
                Q(user__post__unit=request.user.post.unit) | Q(members__in=[request.user])).distinct()
        else:
            session_list = Session.objects.filter(Q(user=request.user) | Q(members__in=[request.user]))
        approval_list = Approval.objects.filter(Q(session__in=session_list) | Q(members__in=[request.user])).distinct()
        _status = request.GET.get('status', 'همه')
        _user = request.GET.get('user', None)
        session = request.GET.get('session', None)
        q = request.GET.get('q', None)
        if _status == 'انجام شده':
            approval_list = approval_list.filter(is_done=True)
        elif _status == 'دارای تأخیر':
            approval_list = approval_list.filter(is_done=False, deadline__isnull=False,
                                                 deadline__lte=jdatetime.date.today())
        elif _status == 'دارای مهلت':
            approval_list = approval_list.filter(is_done=False).filter(
                Q(deadline=None) | Q(deadline__gt=jdatetime.date.today()))
        if _user:
            user_list = _user.split(' ')
            user_filter = reduce(operator.and_, (
                Q(members__first_name__icontains=user_text) | Q(members__last_name__icontains=user_text) | Q(
                    members__personnel_code__icontains=user_text) for user_text in user_list))
            approval_list = approval_list.filter(user_filter)
        if q:
            q_list = q.split(' ')
            q_filter = reduce(operator.and_, (Q(title__icontains=q_text) for q_text in q_list))
            approval_list = approval_list.filter(q_filter)
        if session:
            approval_list = approval_list.filter(session__title__contains=session)
        size = request.user.profile.page_size
        page = min(int(request.GET.get('page', 1)), approval_list.count() // size + 1)
        data = {
            'count': approval_list.count(),
            'page': page,
            'size': size,
            'list': SerApproval(instance=approval_list[size * (page - 1):size * page], many=True,
                                context={'user': request.user}).data
        }
        return Response(data=data)


class ServantList(ListAPIView):
    serializer_class = SerUserList
    queryset = User.objects.filter(post__level__in=['کاردان', 'خدمات'])

    def get_queryset(self):
        if self.request.user.groups.filter(name='room_catering').exists():
            return User.objects.filter(post__level__in=['کاردان', 'خدمات'])
        return User.objects.none()


class OrderSessionCatering(APIView):
    permission_classes = (IsRoomCateringAdmin,)

    def post(self, request):
        session = get_object_or_404(Session, id=request.data['id'])
        if request.data['need_manager'] and session.manager_accept == 'نامشخص':
            session.need_manager = True
        elif request.data['need_deputy'] and session.deputy_accept == 'نامشخص':
            session.need_deputy = True
        else:
            session.breakfast = request.data['breakfast']
            session.lunch = request.data['lunch']
            session.catering = request.data['catering']
            session.catering_agents.set(request.data['catering_agents'])
            session.order_time = jdatetime.datetime.now()
        session.save()
        return Response(data=SerSessionList(session).data)



class MyRoomList(ListAPIView):
    serializer_class = SerRoom

    def get_queryset(self):
        return Room.objects.filter(Q(public=True) | Q(posts__in=[self.request.user.post])).distinct()


class OverlapCheck(GenericAPIView):
    def post(self, request):
        room = get_object_or_404(Room, id=request.data['room'])
        session_list = Session.objects.filter(room=room, date=request.data['date']).exclude(
            id=request.data['id']).exclude(accept_room=False)
        if room.public:
            start = (datetime.datetime.strptime(request.data['start'], '%H:%M') - datetime.timedelta(
                minutes=30)).strftime('%H:%M')
            end = (datetime.datetime.strptime(request.data['end'], '%H:%M') + datetime.timedelta(minutes=30)).strftime(
                '%H:%M')
        else:
            start = request.data['start']
            end = request.data['end']
        data = {
            'ok': not session_list.filter(end__gt=start, start__lt=end).exists(),
            'list': [{
                'title': s.title if s.user == request.user or request.user in s.members.all() else '-----',
                'unit': s.user.post.unit.title,
                'start': str(s.start)[:5],
                'end': str(s.end)[:5]
            } for s in session_list]
        }
        return Response(data=data)


class PublicRoomRequestList(ListAPIView):
    serializer_class = SerRoomRequest

    def get_queryset(self):
        if self.request.user.groups.filter(name='room_admin').exists():
            return apps.get_model('pm.Session').objects.filter(room__public=True, accept_room=None)
        return apps.get_model('pm.Session').objects.filter(id=0)


class PublicRoomRequestAccept(GenericAPIView):
    def post(self, request):
        if not request.user.groups.filter(name='room_admin').exists():
            return Response(data='شما دسترسی لازم ندارید', status=status.HTTP_403_FORBIDDEN)
        session = get_object_or_404(Session, id=request.data['id'], room__public=True, accept_room=None)
        session.accept_room = request.data['accept_room']
        if session.accept_room:
            session.accept_photography = request.data['accept_photography'] or False
            session.accept_filming = request.data['accept_filming'] or False
            session.accept_recording = request.data['accept_recording'] or False
            session.accept_news = request.data['accept_news'] or False
            session.accept_presentation = request.data['accept_presentation'] or False
            session.room_agents.set(request.data['room_agents'])
            session.photography_agents.set(request.data['photography_agents'])
            session.filming_agents.set(request.data['filming_agents'])
            session.recording_agents.set(request.data['recording_agents'])
            session.news_agents.set(request.data['news_agents'])
            session.presentation_agents.set(request.data['presentation_agents'])
        else:
            session.accept_photography = None
            session.accept_filming = None
            session.accept_recording = None
            session.accept_news = None
            session.accept_presentation = None
            session.room_agents.set([])
            session.photography_agents.set([])
            session.filming_agents.set([])
            session.recording_agents.set([])
            session.news_agents.set([])
            session.presentation_agents.set([])
        session.save()
        # ایجاد وظایف
        if session.accept_room:
            for agent in session.room_agents.all():
                job = Job.objects.create(session=session, title='هماهنگی سالن جلسه',
                                         note=f'{session.title}\n{session.room.title}\n{str(session.date)}\n{str(session.start)[:-3]}',
                                         deadline=session.date)
                job.tasks.create(user=request.user, is_owner=True, is_committed=agent == request.user)
                if agent != request.user:
                    job.tasks.create(user=agent)
        if session.accept_photography:
            for agent in session.photography_agents.all():
                job = Job.objects.create(session=session, title='عکسبرداری جلسه',
                                         note=f'{session.title}\n{session.room.title}\n{str(session.date)}\n{str(session.start)[:-3]}',
                                         deadline=session.date)
                job.tasks.create(user=request.user, is_owner=True, is_committed=agent == request.user)
                if agent != request.user:
                    job.tasks.create(user=agent)
        if session.accept_filming:
            for agent in session.filming_agents.all():
                job = Job.objects.create(session=session, title='تصویربرداری جلسه',
                                         note=f'{session.title}\n{session.room.title}\n{str(session.date)}\n{str(session.start)[:-3]}',
                                         deadline=session.date)
                job.tasks.create(user=request.user, is_owner=True, is_committed=agent == request.user)
                if agent != request.user:
                    job.tasks.create(user=agent)
        if session.accept_recording:
            for agent in session.recording_agents.all():
                job = Job.objects.create(session=session, title='ضبط جلسه',
                                         note=f'{session.title}\n{session.room.title}\n{str(session.date)}\n{str(session.start)[:-3]}',
                                         deadline=session.date)
                job.tasks.create(user=request.user, is_owner=True, is_committed=agent == request.user)
                if agent != request.user:
                    job.tasks.create(user=agent)
        if session.accept_news:
            for agent in session.news_agents.all():
                job = Job.objects.create(session=session, title='تهیه خبر جلسه',
                                         note=f'{session.title}\n{session.room.title}\n{str(session.date)}\n{str(session.start)[:-3]}',
                                         deadline=session.date)
                job.tasks.create(user=request.user, is_owner=True, is_committed=agent == request.user)
                if agent != request.user:
                    job.tasks.create(user=agent)
        if session.accept_presentation:
            for agent in session.presentation_agents.all():
                job = Job.objects.create(session=session, title='هماهنگی ارائه جلسه',
                                         note=f'{session.title}\n{session.room.title}\n{str(session.date)}\n{str(session.start)[:-3]}',
                                         deadline=session.date)
                job.tasks.create(user=request.user, is_owner=True, is_committed=False)
                if agent != request.user:
                    job.tasks.create(user=agent)
        return Response(data=request.user.todo_calendar)

# Flow:


def get_node_list(request):
    node_list = request.user.nodes.order_by('-id')
    _type = request.GET.get('type', 'همه')
    if _type != 'همه':
        node_list = node_list.filter(done_time__isnull=_type == 'منتظر اقدام')
    flow = int(request.GET.get('flow', 0))
    if flow:
        node_list = node_list.filter(flow__flow_pattern_id=flow)
    q = request.GET.get('q', None)
    if q:
        q_list = q.split(' ')
        q_filter = reduce(operator.and_, (
        Q(flow__flow_pattern__title__contains=q_text) | Q(flow__user__personnel_code__contains=q_text) | Q(
            flow__user__first_name__contains=q_text) | Q(flow__user__last_name__contains=q_text) | Q(
            flow__user__post__unit__title__contains=q_text) | Q(flow__answers__body__contains=q_text) | Q(
            flow_id=q_text if q_text.isdigit() else 0) for q_text in q_list))
        node_list = node_list.filter(q_filter).distinct()
    return node_list


class NodeList(GenericAPIView):
    def get(self, request):
        node_list = get_node_list(request)
        size = request.user.profile.page_size
        page = min(int(request.GET.get('page', 1)), node_list.count() // 10 + 1)
        data = {
            'type': request.GET.get('type', 'همه'),
            'q': request.GET.get('q', None),
            'count': node_list.count(),
            'page': page,
            'size': size,
            'flow': int(request.GET.get('flow', 0)),
            'list': SerNode(node_list[size * (page - 1):size * page], many=True).data
        }
        return Response(data=data)


class NodeListExcel(GenericAPIView):
    def get(self, request):
        node_list = get_node_list(request)
        wb = Workbook()
        ws = wb.active
        ws.title = f'flows-{str(jdatetime.date.today())}'
        ws.sheet_view.rightToLeft = True
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 35
        ws.column_dimensions['H'].width = 16
        ws.column_dimensions['I'].width = 16
        ws.append(['گزارش:', 'فهرست فرآیندها', '', 'توسط:', request.user.get_full_name(), '', '', 'زمان:',
                   jdatetime.datetime.now().strftime('%Y-%m-%d %H:%M')])
        ws.append(['', '', '', '', '', '', '', '', ''])
        ws.append(['شماره', 'وضعیت', 'فرآیند', 'گره', 'متقاضی', 'کد پرسنلی', 'واحد', 'زمان دریافت', 'زمان ارسال'])
        for node in node_list:
            ws.append([node.flow_id,
                       'مشاهده نشده' if node.seen_time is None else 'منتظر اقدام' if node.done_time is None else 'اقدام شده',
                       node.flow.flow_pattern.title, node.node_pattern.title, node.flow.user.get_full_name(),
                       node.flow.user.personnel_code, node.flow.user.post.unit.title if node.flow.user.post else '',
                       node.create_time.strftime('%Y-%m-%d %H:%M'),
                       node.done_time.strftime('%Y-%m-%d %H:%M') if node.done_time else ''])
        table = Table(displayName="FlowTable", ref=f'A3:I{ws.max_row}')
        table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium9', showRowStripes=True)
        ws.add_table(table)
        farsi_style = styles.NamedStyle(name='farsi_style')
        farsi_style.font = styles.Font(name='Sahel', size=10)
        wb.add_named_style(farsi_style)
        for row in ws.iter_rows():
            for cell in row:
                cell.style = 'farsi_style'
        grey_style = styles.NamedStyle(name='grey_style')
        grey_style.font = styles.Font(name='Sahel', size=10, color='FF808080')
        grey_style.alignment = styles.Alignment(horizontal='left', vertical='center')
        for i in ['A1', 'D1', 'H1']:
            cell = ws[i]
            cell.style = grey_style
        bold_style = styles.NamedStyle(name='bold_style')
        bold_style.font = styles.Font(name='Sahel', size=10, bold=True)
        bold_style.fill = styles.PatternFill(start_color="FFFFE4E1", end_color="FFFFE4E1", fill_type="solid")
        for i in ['B1', 'E1', 'I1']:
            cell = ws[i]
            cell.style = bold_style
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(content=output.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=flows-{str(jdatetime.date.today())}.xlsx'
        return response


class MyFlowPatterList(GenericAPIView):
    def get(self, request):
        data = [{'id': fp.id, 'title': fp.title, 'type': fp.flow_type.title if fp.flow_type else ''} for fp in
                FlowPattern.objects.filter(posts__in=[request.user.post], active=True).order_by('title') if
                fp.quota_per_user > request.user.flows.filter(flow_pattern=fp).count()]
        return Response(data=data)


class AllFlowPatterList(GenericAPIView):
    def get(self, request):
        data = [{'id': fp.id, 'title': fp.title} for fp in
                FlowPattern.objects.filter(active=True, nodes__nodes__user=request.user).distinct()]
        return Response(data=data)


def get_field_answer(field, flow, order=0):
    answer = field.answers.filter(flow=flow, order=order).first()
    return answer.body if answer and answer.body else None


def get_field_file(field, flow, order=0):
    answer = field.answers.filter(flow=flow, order=order).first()
    return str(answer.file) if answer and answer.file else None


class NodeDetail(GenericAPIView):
    def get(self, request, pk):
        node = get_object_or_404(Node, pk=pk, user=request.user)
        if node.seen_time is None:
            node.seen_time = jdatetime.datetime.now()
            node.save()
        request.user.notifications.filter(node=node).update(seen_time=jdatetime.datetime.now())
        flow = node.flow
        data = SerNode(node).data
        data['fields'] = []
        for nodefield in node.node_pattern.fields.all():
            if nodefield.field.table:
                if nodefield.field.table not in map(lambda a: a['label'], data['fields']):
                    head = []
                    for f in node.node_pattern.fields.filter(field__table=nodefield.field.table):
                        head.append({
                            'id': f.field.id,
                            'label': f.field.label,
                            'hint': f.field.hint,
                            'type': f.field.type,
                            'choices': f.field.choices,
                            'answer': '',
                            'file': None,
                            'new_file': None,
                            'editable': f.editable,
                            'required': f.required
                        })
                    rows = []
                    for i in range(nodefield.field.answers.filter(flow=flow).count()):
                        row = []
                        for f in node.node_pattern.fields.filter(field__table=nodefield.field.table):
                            row.append({
                                'id': f.field.id,
                                'label': f.field.label,
                                'hint': f.field.hint,
                                'type': f.field.type,
                                'choices': f.field.choices,
                                'answer': get_field_answer(field=f.field, flow=flow, order=i),
                                'file': get_field_file(field=f.field, flow=flow, order=i),
                                'new_file': None,
                                'editable': f.editable,
                                'required': f.required
                            })
                        rows.append(row)
                    data['fields'].append({
                        'id': 0,
                        'label': nodefield.field.table,
                        'type': 'table',
                        'row_min': nodefield.field.row_min,
                        'row_max': nodefield.field.row_max,
                        'head': head,
                        'rows': rows,
                    })
            else:
                data['fields'].append({
                    'id': nodefield.field.id,
                    'label': nodefield.field.label,
                    'hint': nodefield.field.hint,
                    'type': nodefield.field.type,
                    'choices': nodefield.field.choices,
                    'table': nodefield.field.table,
                    'answer': get_field_answer(field=nodefield.field, flow=flow),
                    'file': get_field_file(field=nodefield.field, flow=flow),
                    'new_file': None,
                    'editable': nodefield.editable,
                    'required': nodefield.required
                })
        return Response(data=data)


class GetNodePdf(GenericAPIView):
    def get(self, request, pk):
        node = get_object_or_404(Node, pk=pk, user=request.user)
        html_string = render_to_string('node.html', {'node': node})
        response = HttpResponse(HTML(string=html_string).write_pdf(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="cover{node.flow_id}.pdf"'
        return response


def check_ifs(dispatch, node):
    if not dispatch.ifs.exists():
        return True
    if dispatch.if_operator == 'and':
        for item in dispatch.ifs.all():
            answer = get_field_answer(field=item.key, flow=node.flow, order=0)
            file = get_field_file(field=item.key, flow=node.flow, order=0)
            if (item.type == 'خالی' and ((item.key.type != 'file' and answer) or (item.key.type == 'file' and file))) \
                    or (item.type == 'دارای مقدار' and (
                    (item.key.type != 'file' and answer is None) or (item.key.type == 'file' and file is None))) \
                    or (item.type == 'مساوی' and answer != item.value) \
                    or (item.type == 'نامساوی' and answer == item.value) \
                    or (item.type == 'بزرگتر' and answer <= item.value) \
                    or (item.type == 'بزرگتر یا مساوی' and answer < item.value) \
                    or (item.type == 'کوچکتر' and answer >= item.value) \
                    or (item.type == 'کوچکتر یا مساوی' and answer > item.value):
                return False
            # وقتی نوع فیلد «چندانتخابی» است آرایه گزینه‌های فیلد و آرایه گزینه‌های شرط اگر اشتراک داشت جواب مثبت است
            if item.type == 'موجود در لیست':
                if item.key.type == 'multi-select':
                    if not bool(set(answer.split(',')) & set(item.values)):
                        return False
                elif answer not in item.values:
                    return False
            if item.type == 'ناموجود در لیست':
                if item.key.type == 'multi-select':
                    if bool(set(answer.split(',')) & set(item.values)):
                        return False
                elif answer in item.values:
                    return False
        return True
    else:
        for item in dispatch.ifs.all():
            answer = get_field_answer(field=item.key, flow=node.flow, order=0)
            file = get_field_file(field=item.key, flow=node.flow, order=0)
            if (item.type == 'خالی' and (
                    (item.key.type != 'file' and answer is None) or (item.key.type == 'file' and file is None))) \
                    or (item.type == 'دارای مقدار' and (
                    (item.key.type != 'file' and answer) or (item.key.type == 'file' and file))) \
                    or (item.type == 'مساوی' and answer == item.value) \
                    or (item.type == 'نامساوی' and answer != item.value) \
                    or (item.type == 'بزرگتر' and answer > item.value) \
                    or (item.type == 'بزرگتر یا مساوی' and answer >= item.value) \
                    or (item.type == 'کوچکتر' and answer < item.value) \
                    or (item.type == 'کوچکتر یا مساوی' and answer <= item.value):
                return True
            # وقتی نوع فیلد «چندانتخابی» است آرایه گزینه‌های فیلد و آرایه گزینه‌های شرط اگر اشتراک داشت جواب مثبت است
            if item.type == 'موجود در لیست':
                if item.key.type == 'multi-select':
                    if bool(set(answer.split(',')) & set(item.values)):
                        return True
                elif answer in item.values:
                    return True
            if item.type == 'ناموجود در لیست':
                if item.key.type == 'multi-select':
                    if not bool(set(answer.split(',')) & set(item.values)):
                        return True
                elif answer not in item.values:
                    return True
        return False


class NodeSave(GenericAPIView):
    def post(self, request):
        if request.data['node'] == '0':
            node_pattern = get_object_or_404(NodePattern, pk=request.data['node_pattern'])
            if node_pattern.flow_pattern.quota_per_user <= request.user.flows.filter(
                    flow_pattern=node_pattern.flow_pattern).count():
                return Response(data='سقف تعداد فرآیند تکمیل شده است', status=status.HTTP_400_BAD_REQUEST)
            flow = Flow.objects.create(user=request.user, flow_pattern_id=node_pattern.flow_pattern_id)
            node = Node.objects.create(flow=flow, user=request.user, post=request.user.post, node_pattern=node_pattern,
                                       seen_time=jdatetime.datetime.now())
            field = Field.objects.filter(flow_pattern=node_pattern.flow_pattern, label='level').first()
            if field:
                Answer.objects.create(flow=flow, field=field, body=request.user.post.level)
            field = Field.objects.filter(flow_pattern=node_pattern.flow_pattern, label='unit').first()
            if field:
                Answer.objects.create(flow=flow, field=field, body=request.user.post.unit.title)
        else:
            node = get_object_or_404(Node, pk=request.data['node'], user=request.user)
        for data in request.data.getlist('removing_rows', []):
            table_name = data[:data.rfind('-')]
            row_index = int(data[data.rfind('-') + 1:])
            Answer.objects.filter(flow=node.flow, field__table=table_name, order=row_index).delete()
        for data in request.data.getlist('removing_files', []):
            field_id = int(data[:data.find('-')])
            row_index = int(data[data.find('-') + 1:])
            Answer.objects.filter(flow=node.flow, field_id=field_id, order=row_index).delete()
        for data in request.data:
            if data[:6] == 'answer':
                tmp = data[7:]
                field_id = int(tmp[:tmp.find('-')])
                row_index = int(tmp[tmp.find('-') + 1:])
                answer, created = Answer.objects.get_or_create(flow=node.flow, field_id=field_id, order=row_index)
                answer.body = request.data[data]
                answer.save()
            elif data[:4] == 'file':
                tmp = data[5:]
                field_id = int(tmp[:tmp.find('-')])
                row_index = int(tmp[tmp.find('-') + 1:])
                answer, created = Answer.objects.get_or_create(flow=node.flow, field_id=field_id, order=row_index)
                answer.file = request.data[data]
                answer.save()
        node.done_time = jdatetime.datetime.now()
        node.save()
        owner = node.flow.nodes.order_by('pk')[0].user
        for dispatch in Dispatch.objects.filter(start=node.node_pattern):
            if check_ifs(dispatch=dispatch, node=node):
                if dispatch.send_to_owner:
                    new_node = Node.objects.create(flow=node.flow, user=node.flow.user, post=node.flow.user.post,
                                                   node_pattern=dispatch.end)
                    Notification.objects.create(user_id=node.flow.user_id, title=node.flow.flow_pattern.title,
                                                body=node.node_pattern.title, url=f'/flow?id={new_node.id}',
                                                node=new_node)
                if dispatch.send_to_parent:
                    parent = owner.post.parent.active_user or owner.post.unit.manager or owner.post.unit.parent.manager
                    new_node = Node.objects.create(flow=node.flow, user=parent, post=parent.post,
                                                   node_pattern=dispatch.end)
                    Notification.objects.create(user=parent, title=node.flow.flow_pattern.title,
                                                body=node.node_pattern.title, url=f'/flow?id={new_node.id}',
                                                node=new_node)
                    # اگر کاربر موردنظر یافت نشود به مدیر مجموعه ارجاع می‌شود
                if dispatch.send_to_manager:
                    new_node = Node.objects.create(flow=node.flow, user=owner.post.unit.manager,
                                                   post=owner.post.unit.manager.post if owner.post.unit.manager else None,
                                                   node_pattern=dispatch.end)
                    Notification.objects.create(user=owner.post.unit.manager, title=node.flow.flow_pattern.title,
                                                body=node.node_pattern.title, url=f'/flow?id={new_node.id}',
                                                node=new_node)
                for post in dispatch.send_to_posts.all():
                    user = post.active_user or post.parent.active_user
                    if user:
                        new_node = Node.objects.create(flow=node.flow, user=user, post=user.post,
                                                       node_pattern=dispatch.end)
                        Notification.objects.create(user=user, title=node.flow.flow_pattern.title,
                                                    body=node.node_pattern.title, url=f'/flow?id={new_node.id}',
                                                    node=new_node)
        if node.node_pattern.next:
            for user_id in request.data.getlist('next_users', []):
                user = User.objects.get(id=user_id)
                new_node = Node.objects.create(flow=node.flow, user=user, post=user.post,
                                               node_pattern=node.node_pattern.next)
                Notification.objects.create(user=user, title=node.flow.flow_pattern.title, body=node.node_pattern.title,
                                            url=f'/flow?id={new_node.id}', node=new_node)
        data = {
            'todo_flow': request.user.todo_flow,
            'node': SerNode(node).data,
        }
        return Response(data=data)


class NodeRemove(GenericAPIView):
    def post(self, request):
        node = get_object_or_404(Node, pk=request.data['pk'], user=request.user)
        if node.removable:
            node.flow.delete()
            return Response(data=request.user.todo_task)
        return Response(data='denied', status=status.HTTP_403_FORBIDDEN)


class NodeRevert(GenericAPIView):
    def post(self, request):
        node = get_object_or_404(Node, pk=request.data['pk'], user=request.user, done_time__isnull=False)
        ends = Dispatch.objects.filter(start=node.node_pattern).values_list('end', flat=True)
        if Node.objects.filter(node_pattern_id__in=ends, flow=node.flow, create_time__gte=node.done_time,
                               create_time__lt=(node.done_time + jdatetime.timedelta(seconds=1)),
                               seen_time__isnull=False).exists():
            return Response(data='گره بعدی توسط یکی از کاربران مشاهده شد', status=status.HTTP_403_FORBIDDEN)
        Node.objects.filter(node_pattern_id__in=ends, flow=node.flow, create_time__gte=node.done_time,
                            create_time__lt=(node.done_time + jdatetime.timedelta(seconds=1))).delete()
        node.done_time = None
        node.save()
        fields = NodeField.objects.filter(node=node.node_pattern, editable=True).values_list('field', flat=True)
        Answer.objects.filter(flow=node.flow, field_id__in=fields).delete()
        data = {
            'todo_flow': request.user.todo_flow,
            'node': SerNode(node).data,
        }
        return Response(data=data)


class StartNewFlow(GenericAPIView):
    def get(self, request, pk):
        node_pattern = NodePattern.objects.filter(flow_pattern_id=pk, flow_pattern__posts__in=[request.user.post],
                                                  is_first=True).first()
        data = {'id': 0,
                'form_width': node_pattern.flow_pattern.form_width,
                'quota_per_user': node_pattern.flow_pattern.quota_per_user,
                'flow_title': node_pattern.flow_pattern.title,
                'flow_user_name': request.user.get_full_name(),
                'flow_user_photo': request.user.photo_url,
                'flow_user_department': request.user.post.unit.department.title,
                'flow_user_personnel_code': request.user.personnel_code,
                'node_pattern': node_pattern.id,
                'node_title': node_pattern.title,
                'node_next': node_pattern.next_id,
                'create_time': str(jdatetime.datetime.now()),
                'done_time': None,
                'preamble': node_pattern.flow_pattern.preamble,
                'poster': node_pattern.flow_pattern.poster.name if node_pattern.flow_pattern.poster else None,
                'image': node_pattern.flow_pattern.image.name if node_pattern.flow_pattern.image else None,
                'fields': []
                }
        for nodefield in node_pattern.fields.all():
            if nodefield.field.table:
                if nodefield.field.table not in map(lambda a: a['label'], data['fields']):
                    head = []
                    for f in node_pattern.fields.filter(field__table=nodefield.field.table):
                        head.append({
                            'id': f.field.id,
                            'label': f.field.label,
                            'hint': f.field.hint,
                            'type': f.field.type,
                            'choices': f.field.choices,
                            'answer': [] if f.field.type == 'multi-select' else '',
                            'file': None,
                            'new_file': None,
                            'editable': f.editable,
                            'required': f.required
                        })
                    data['fields'].append({
                        'id': 0,
                        'label': nodefield.field.table,
                        'type': 'table',
                        'row_min': nodefield.field.row_min,
                        'row_max': nodefield.field.row_max,
                        'head': head,
                        'rows': []
                    })
            else:
                data['fields'].append({
                    'id': nodefield.field.id,
                    'label': nodefield.field.label,
                    'hint': nodefield.field.hint,
                    'type': nodefield.field.type,
                    'choices': nodefield.field.choices,
                    'table': nodefield.field.table,
                    'answer': [] if nodefield.field.type == 'multi-select' else '',
                    'file': None,
                    'new_file': None,
                    'editable': nodefield.editable,
                    'required': nodefield.required
                })
        return Response(data=data)


class FlowHistory(GenericAPIView):
    def get(self, request, pk):
        node = get_object_or_404(Node, pk=pk, user=request.user)
        return Response(data=SerFlowHistory(node.flow.nodes, many=True).data)


class FlowPatternAdd(CreateAPIView):
    permission_classes = [IsFlowAdmin]
    queryset = FlowPattern.objects.all()
    serializer_class = SerFlowPatternDetail


class FlowPatternRemove(DestroyAPIView):
    permission_classes = [IsFlowAdmin]
    queryset = FlowPattern.objects.all()


class FlowPatternList(GenericAPIView):
    permission_classes = [IsFlowAdmin]

    def get(self, request):
        _type = request.GET.get('type', 'همه')
        if _type == 'همه':
            flow_pattern_list = FlowPattern.objects.all()
        else:
            flow_pattern_list = FlowPattern.objects.filter(flow_type__title=_type)
        size = request.user.profile.page_size
        page = min(int(request.GET.get('page', 1)), flow_pattern_list.count() // 10 + 1)
        data = {
            'count': flow_pattern_list.count(),
            'page': page,
            'size': size,
            'type': _type,
            'list': SerFlowPatternList(instance=flow_pattern_list[size * (page - 1):size * page], many=True).data
        }
        return Response(data=data)


class FlowPatternDetail(RetrieveAPIView):
    permission_classes = [IsFlowAdmin]
    queryset = FlowPattern.objects.all()
    serializer_class = SerFlowPatternDetail


class FlowPatternFields(ListAPIView):
    permission_classes = [IsFlowAdmin]
    serializer_class = SerField

    def get_queryset(self):
        flow_pattern = get_object_or_404(FlowPattern, pk=self.kwargs['pk'])
        return flow_pattern.fields


class FlowPatternNodes(ListAPIView):
    permission_classes = [IsFlowAdmin]
    serializer_class = SerNodePattern

    def get_queryset(self):
        flow_pattern = get_object_or_404(FlowPattern, pk=self.kwargs['pk'])
        return flow_pattern.nodes


class SaveFlowPatternDetail(GenericAPIView):
    permission_classes = [IsFlowAdmin]

    def post(self, request):
        if request.data['id']:
            fp = get_object_or_404(FlowPattern, pk=request.data['id'])
            fp.title = request.data['title']
            # Set the flow_type based on the type value
            flow_type_obj, created = FlowPatternType.objects.get_or_create(title=request.data['type'])
            fp.flow_type = flow_type_obj
            fp.form_width = request.data['form_width']
            fp.quota_per_user = request.data['quota_per_user']
            fp.active = request.data['active'] == 'true'
            fp.preamble = request.data['preamble']
            fp.save()
        else:
            # Create or get the flow_type object
            flow_type_obj, created = FlowPatternType.objects.get_or_create(title=request.data['type'])
            fp = FlowPattern.objects.create(title=request.data['title'], flow_type=flow_type_obj,
                                            form_width=request.data['form_width'],
                                            quota_per_user=request.data['quota_per_user'],
                                            active=request.data['active'], preamble=request.data['preamble'])
        if 'new_poster' in request.data:
            fp.poster = request.data['new_poster']
        elif request.data['poster'] == 'removed':
            fp.poster = None
        if 'new_image' in request.data:
            fp.image = request.data['new_image']
        elif request.data['image'] == 'removed':
            fp.image = None
        fp.save()
        fp.posts.set(request.data.getlist('posts'))
        return Response(data=SerFlowPatternDetail(fp).data)


class SaveFlowPatternFields(GenericAPIView):
    permission_classes = [IsFlowAdmin]

    def post(self, request):
        flow_pattern = get_object_or_404(FlowPattern, id=request.data['id'])
        ids = list(map(lambda f: f['id'], request.data['fields']))
        flow_pattern.fields.exclude(id__in=ids).delete()
        for data_index, data in enumerate(request.data['fields']):
            if data['id'] == 0:
                flow_pattern.fields.create(label=data['label'], hint=data['hint'], type=data['type'],
                                           choices=data['choices'], table=data['table'], row_min=data['row_min'],
                                           row_max=data['row_max'], order=data_index, is_archived=data['is_archived'])
            else:
                flow_pattern.fields.filter(id=data['id']).update(label=data['label'], hint=data['hint'],
                                                                 type=data['type'], choices=data['choices'],
                                                                 table=data['table'], row_min=data['row_min'],
                                                                 row_max=data['row_max'], order=data_index,
                                                                 is_archived=data['is_archived'])
        return Response(data=SerField(flow_pattern.fields, many=True).data)


class SaveFlowPatternNodes(GenericAPIView):
    permission_classes = [IsFlowAdmin]

    def post(self, request):
        flow_pattern = get_object_or_404(FlowPattern, id=request.data['id'])
        # remove unlisted nodes:
        ids = list(map(lambda f: f['id'], request.data['nodes']))
        flow_pattern.nodes.exclude(id__in=ids).delete()
        # create or update nodes:
        for data_index, data in enumerate(request.data['nodes']):
            if data['id'] == 0:
                node = flow_pattern.nodes.create(title=data['title'], order=data_index, next_id=data['next'],
                                                 is_archived=data['is_archived'], is_bottleneck=data['is_bottleneck'],
                                                 respite=data['respite'])
            else:
                node = flow_pattern.nodes.get(id=data['id'])
                node.title = data['title']
                node.order = data_index
                node.next_id = data['next']
                node.is_archived = data['is_archived']
                node.is_bottleneck = data['is_bottleneck']
                node.is_first = data['is_first']
                node.sms = data['sms']
                node.respite = data['respite']
                node.save()
            # remove unlisted node_fields:
            ids = list(map(lambda f: f['field'], data['fields']))
            node.fields.exclude(field_id__in=ids).delete()
            # create or update node_fields:
            for field_data in data['fields']:
                node_field, created = node.fields.get_or_create(field_id=field_data['field'])
                node_field.editable = field_data['editable']
                node_field.required = field_data['required']
                node_field.save()
            # remove unlisted dispatches:
            ids = list(map(lambda f: f['id'], data['dispatches']))
            Dispatch.objects.filter(start=node).exclude(id__in=ids).delete()
            # create or update dispatches:
            for dispatch_date in data['dispatches']:
                if dispatch_date['id'] == 0:
                    dispatch = Dispatch.objects.create(start=node, end_id=dispatch_date['end'])
                else:
                    dispatch = Dispatch.objects.get(id=dispatch_date['id'])
                dispatch.end_id = dispatch_date['end']
                dispatch.send_to_owner = dispatch_date['send_to_owner']
                dispatch.send_to_parent = dispatch_date['send_to_parent']
                dispatch.send_to_manager = dispatch_date['send_to_manager']
                dispatch.send_to_posts.set(dispatch_date['send_to_posts'])
                dispatch.if_operator = dispatch_date['if_operator']
                dispatch.save()
                ids = list(map(lambda f: f['id'], dispatch_date['ifs']))
                dispatch.ifs.exclude(id__in=ids).delete()
                for dispatch_if_data in dispatch_date['ifs']:
                    if dispatch_if_data['id'] == 0:
                        dispatch_if = DispatchIf.objects.create(dispatch=dispatch, key_id=dispatch_if_data['key'],
                                                                type=dispatch_if_data['type'])
                    else:
                        dispatch_if = DispatchIf.objects.get(pk=dispatch_if_data['id'])
                        dispatch_if.key_id = dispatch_if_data['key']
                        dispatch_if.type = dispatch_if_data['type']
                    dispatch_if.value = dispatch_if_data['value']
                    dispatch_if.values = dispatch_if_data['values']
                    dispatch_if.save()
        return Response(data=SerNodePattern(flow_pattern.nodes, many=True).data)


class PostListForFlowPatternManagement(ListAPIView):
    permission_classes = [IsFlowAdmin]
    serializer_class = SerPostList
    queryset = Post.objects.all()



class FlowCategoryAPI(APIView):
    def get(self, request):
        resp = []
        # Get all flow pattern types from the FlowPatternType model
        flow_type_objects = FlowPatternType.objects.all()
        for flow_type_obj in flow_type_objects:
            items = FlowPattern.objects.filter(posts__in=[request.user.post], flow_type=flow_type_obj, active=True)

            # Handle special case for financial resources
            if flow_type_obj.title == 'مالی و سرمایه‌های انسانی':
                financial_type_obj, created = FlowPatternType.objects.get_or_create(title='مالی و سرمایه انسانی')
                items = FlowPattern.objects.filter(posts__in=[request.user.post], flow_type__in=[flow_type_obj, financial_type_obj], active=True)

            data = []
            for item in items:
                data.append({
                    'id': item.pk,
                    'title': item.title,
                    'is_active': item.active,
                })
            resp.append({
                'name': flow_type_obj.title,
                'count': len(items),
                'items': data,
            })

        return Response(resp)



class SessionAddOrUpdateRate(APIView):
    def post(self, request):
        user = request.user
        rate_num = request.data.get('rate')
        session_id = request.data.get('id')

        session = get_object_or_404(Session, pk=session_id)
        rate = SessionRate.objects.filter(user=user, session=session).first()
        if not rate:
            rate = SessionRate()
        rate.session = session
        rate.rate = rate_num
        rate.user = user
        rate.save()

        return Response({"rate": rate_num, "session_id": session_id})



from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django.utils import timezone
import random


# -----------------------------
# Pagination اختصاصی هر ستون
# -----------------------------
class ColumnPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'


# -----------------------------
# تولید دیتای فیک تسک
# -----------------------------
def generate_fake_tasks(tag_id, status, total=75):
    tasks = []

    for i in range(total):
        tasks.append({
            "id": 10000 + i,
            "order": i,
            "is_seen": random.choice([True, False]),
            "unseen_chat_count": random.randint(0, 3),
            "job": {
                "id": 20000 + i,
                "title": f"تسک نمونه {i} - {status}",
                "deadline": "1404-12-01",
                "urgency": random.randint(1, 3),
                "status": status,
                "respite": random.randint(-10, 15),
                "project": 637,
                "has_chat": random.choice([True, False]),
                "is_owner": True,
                "is_informees": False,
                "assignees": [
                    {
                        "user": 600,
                        "name": "محمدرضا صالحی",
                        "photo_url": None,
                        "is_owner": True
                    }
                ]
            }
        })

    return tasks


# -----------------------------
# ویو اصلی تستی
# -----------------------------
class TaskListV2Test(APIView):

    def get(self, request):

        tag = request.GET.get("tag")
        status_param = request.GET.get("status")

        # ---------------------------------
        # اگر درخواست pagination یک ستون بود
        # ---------------------------------
        if tag is not None and status_param is not None:
            total_count = 75  # فرضی
            tasks = generate_fake_tasks(tag, status_param, total_count)

            paginator = ColumnPagination()
            page = paginator.paginate_queryset(tasks, request)

            return paginator.get_paginated_response(page)

        # ---------------------------------
        # در غیر این صورت کل board را بده
        # ---------------------------------
        columns = []

        fake_tags = [
            {"id": None, "title": "بدون برچسب"},
            {"id": 373, "title": "پرتال توسعه"},
            {"id": 400, "title": "اپ موبایل"},
        ]

        statuses = ["todo", "doing", "done"]

        for tag_item in fake_tags:
            tag_id = tag_item["id"]

            column_statuses = {}

            for st in statuses:
                total_count = random.randint(15, 120)
                tasks = generate_fake_tasks(tag_id, st, total_count)

                paginator = ColumnPagination()
                page = paginator.paginate_queryset(tasks, request)

                column_statuses[st] = {
                    "count": total_count,
                    "next": (
                        f"/pm/task-list/v2/?tag={tag_id}&status={st}&page=2"
                        if total_count > paginator.page_size else None
                    ),
                    "previous": None,
                    "results": page
                }

            columns.append({
                "id": tag_id,
                "title": tag_item["title"],
                "statuses": column_statuses
            })

        return Response({
            "meta": {
                "archive": False,
                "team": True,
                "page_size": 20,
                "generated_at": timezone.now()
            },
            "columns": columns
        })
